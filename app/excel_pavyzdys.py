import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell
from io import BytesIO
from datetime import datetime, timedelta
import json
import shutil
import os

def generate_pavyzdys_excel_report(campaign_id: int):
    """Generate Excel report by copying pavyzdys1.xlsx and inserting campaign data"""
    # Import needed functions 
    from app.models import get_campaign_report_data, load_trp_distribution
    
    data = get_campaign_report_data(campaign_id)
    if not data:
        return None
    
    campaign = data['campaign']
    waves = data['waves']
    
    # Load TRP calendar data
    trp_data = load_trp_distribution(campaign_id)
    
    # Path to the template file
    template_path = "/home/vainiusl/py_projects/tv-planner/pavyzdys1.xlsx"
    
    # Load the template Excel file
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
    except Exception as e:
        print(f"Error loading template file: {e}")
        return None
    
    # Fill in campaign data into the existing template
    # The template already has the structure, we just need to fill the data
    
    # Fill header information (the template already has labels, we fill values)
    ws['C1'] = campaign.get('agency', 'BPN LT')  # Agentūra
    ws['C2'] = campaign.get('client', '')        # Klientas  
    ws['C3'] = campaign.get('product', '')       # Produktas
    ws['C4'] = campaign.get('name', '')          # Kampanija
    ws['C6'] = campaign.get('country', 'Lietuva') # Šalis
    
    # Format period from dates
    if campaign.get('start_date') and campaign.get('end_date'):
        try:
            start = datetime.strptime(campaign['start_date'], '%Y-%m-%d')
            period = start.strftime('%Y.%m')
        except:
            period = campaign.get('start_date', '')[:7]  # YYYY-MM
        ws['C5'] = period  # Periodas
    
    # Get first TVC name and duration if available
    if waves and waves[0]['items']:
        first_item = waves[0]['items'][0]
        # TVC name (assuming it goes in G3 based on previous layout)
        first_tvc = first_item.get('tvc_name', '')
        if 'G3' in [cell.coordinate for row in ws.iter_rows() for cell in row if cell.value]:
            ws['G3'] = first_tvc
            
        # Clip duration (assuming it goes in F5)
        first_duration = first_item.get('tvc_duration', first_item.get('clip_duration', 10))
        if 'F5' in [cell.coordinate for row in ws.iter_rows() for cell in row if cell.value]:
            ws['F5'] = first_duration
    
    # Clear existing data rows (from row 14 onwards) and insert campaign data
    # First, clear old data (but keep the structure) - be careful with merged cells
    max_data_row = 50  # Clear a reasonable range
    for row_num in range(14, max_data_row):
        for col_num in range(1, 30):  # Clear main data columns
            try:
                cell = ws.cell(row=row_num, column=col_num)
                if not isinstance(cell, MergedCell):
                    cell.value = None
            except:
                pass  # Skip cells that can't be cleared
    
    # Update the client name in the header (K10)
    client_name = campaign.get('client', 'Kliento')[:15]
    if client_name:
        ws['K10'] = client_name
    
    # Insert campaign data starting from row 14
    current_row = 14
    total_cost = 0
    
    for wave in waves:
        for item in wave['items']:
            # Calculate values
            channel_share = item.get('channel_share', 0.75)
            pt_zone_share = item.get('pt_zone_share', 0.55)
            
            affinity1 = item.get('affinity1', 88.2)  # Default from template
            affinity2 = item.get('affinity2', 88.2)
            affinity3 = item.get('affinity3', 88.2) 
            trps = item.get('trps', 0)
            grp_planned = (trps * 100 / affinity1) if affinity1 > 0 else trps
            
            # Prices and discounts
            gross_cpp = item.get('gross_cpp_eur', item.get('price_per_sec_no_discount', 18.4))
            
            # Indices
            duration_idx = item.get('duration_index', 1.0)
            seasonal_idx = item.get('seasonal_index', 1.0)
            trp_purchase_idx = item.get('trp_purchase_index', 0.95)
            advance_idx = item.get('advance_purchase_index', 0.95)
            position_idx = item.get('position_index', 1.0)
            
            # Calculate gross price
            clip_duration = item.get('tvc_duration', item.get('clip_duration', 10))
            gross_price = trps * gross_cpp * duration_idx * seasonal_idx * trp_purchase_idx * advance_idx * position_idx
            
            # Discounts
            client_discount = item.get('client_discount', 0)
            net_price = gross_price * (1 - client_discount / 100)
            agency_discount = item.get('agency_discount', 0) 
            net_net_price = net_price * (1 - agency_discount / 100)
            
            total_cost += net_net_price
            
            # Fill data matching the template structure
            ws.cell(row=current_row, column=1).value = item.get('owner', item.get('channel_group', ''))  # A: Kanalas
            ws.cell(row=current_row, column=2).value = item.get('target_group', '')  # B: Perkama TG
            ws.cell(row=current_row, column=3).value = channel_share  # C: kanalo dalis
            ws.cell(row=current_row, column=4).value = pt_zone_share  # D: PT zonos dalis
            ws.cell(row=current_row, column=5).value = f"=$F$5"  # E: trukmė reference
            ws.cell(row=current_row, column=6).value = grp_planned  # F: GRP planuojami
            ws.cell(row=current_row, column=7).value = trps  # G: TRP perkami
            ws.cell(row=current_row, column=8).value = f"=G{current_row}/COUNT(AB{current_row}:XFD{current_row})"  # H: DIENOS TRP1
            # Columns I, J (TRP2, DIENOS TRP2) - leave empty as in template
            ws.cell(row=current_row, column=11).value = grp_planned  # K: Kliento TRP
            ws.cell(row=current_row, column=12).value = affinity1  # L: Affinity1
            ws.cell(row=current_row, column=13).value = affinity2  # M: Affinity2
            ws.cell(row=current_row, column=14).value = affinity3  # N: Affinity3
            
            # Add additional calculated fields if they exist in the template
            if gross_cpp:
                ws.cell(row=current_row, column=15).value = gross_cpp  # O: Gross CPP
            if gross_price:
                ws.cell(row=current_row, column=21).value = gross_price  # U: Gross kaina
            if client_discount:
                ws.cell(row=current_row, column=22).value = client_discount / 100  # V: Kliento nuolaida
            if net_price:
                ws.cell(row=current_row, column=23).value = net_price  # W: Net kaina
            if agency_discount:
                ws.cell(row=current_row, column=24).value = agency_discount / 100  # X: Agentūros nuolaida
            if net_net_price:
                ws.cell(row=current_row, column=25).value = net_net_price  # Y: Net net kaina
            
            current_row += 1
    
    # Add TRP calendar data if available (the template should already have calendar structure)
    if trp_data and isinstance(trp_data, str):
        trp_data = json.loads(trp_data)
    
    # If there's TRP data, fill it into the existing calendar structure
    if trp_data and campaign.get('start_date') and campaign.get('end_date'):
        try:
            start_date = datetime.strptime(campaign['start_date'], '%Y-%m-%d')
            end_date = datetime.strptime(campaign['end_date'], '%Y-%m-%d')
            
            # Calendar starts from column AB (28) - use existing template structure
            cal_start_col = 28
            current_date = start_date
            cal_col = cal_start_col
            
            # Fill TRP values into existing calendar
            while current_date <= end_date and cal_col < 76:  # Max columns
                date_key = current_date.strftime('%Y-%m-%d')
                trp_value = trp_data.get(date_key, 0)
                
                if trp_value > 0:
                    # Fill TRP values for data rows
                    data_row = 14
                    num_items = sum(len(wave['items']) for wave in waves)
                    if num_items > 0:
                        item_trp = trp_value / num_items
                        for wave in waves:
                            for item in wave['items']:
                                ws.cell(row=data_row, column=cal_col).value = item_trp
                                data_row += 1
                
                current_date += timedelta(days=1)
                cal_col += 1
                
        except Exception as e:
            pass  # Skip calendar if date parsing fails
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output