# app/models.py
import sqlite3, os
DB_PATH = os.path.join(os.path.dirname(__file__), "tv-calc.db")

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_db():
    with get_db() as db:
        # -------- Channel groups & channels --------
        db.execute("""
        CREATE TABLE IF NOT EXISTS channel_groups (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        )""")
        db.execute("""
        CREATE TABLE IF NOT EXISTS channels (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            channel_group_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            size TEXT NOT NULL CHECK(size IN ('big','small')),
            UNIQUE(channel_group_id, name),
            FOREIGN KEY(channel_group_id) REFERENCES channel_groups(id) ON DELETE CASCADE
        )""")

        # -------- TRP rates (legacy admin) --------
        db.execute("""
        CREATE TABLE IF NOT EXISTS trp_rates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            owner TEXT NOT NULL,
            target_group TEXT NOT NULL,
            primary_label TEXT NOT NULL,
            secondary_label TEXT,
            share_primary REAL,
            share_secondary REAL,
            prime_share_primary REAL,
            prime_share_secondary REAL,
            price_per_sec_eur REAL NOT NULL,
            UNIQUE(owner, target_group)
        )""")

        # -------- Pricing lists (rate cards) --------
        db.execute("""
        CREATE TABLE IF NOT EXISTS pricing_lists (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        )""")
        # Check if we need to migrate the pricing_list_items table to remove unique constraint
        cursor = db.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='pricing_list_items'")
        table_def = cursor.fetchone()
        
        if table_def and 'UNIQUE(pricing_list_id, owner, target_group)' in table_def['sql']:
            # Table has the old unique constraint, need to migrate
            db.execute("ALTER TABLE pricing_list_items RENAME TO pricing_list_items_old")
            
            # Create new table without unique constraint
            db.execute("""
            CREATE TABLE pricing_list_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pricing_list_id INTEGER NOT NULL,
                owner TEXT NOT NULL,
                target_group TEXT NOT NULL,
                primary_label TEXT NOT NULL,
                secondary_label TEXT,
                share_primary REAL,
                share_secondary REAL,
                prime_share_primary REAL,
                prime_share_secondary REAL,
                price_per_sec_eur REAL NOT NULL,
                FOREIGN KEY(pricing_list_id) REFERENCES pricing_lists(id) ON DELETE CASCADE
            )""")
            
            # Copy data from old table
            db.execute("""
            INSERT INTO pricing_list_items 
            SELECT * FROM pricing_list_items_old
            """)
            
            # Drop old table
            db.execute("DROP TABLE pricing_list_items_old")
        else:
            # Create table without unique constraint if it doesn't exist
            db.execute("""
            CREATE TABLE IF NOT EXISTS pricing_list_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pricing_list_id INTEGER NOT NULL,
                owner TEXT NOT NULL,
                target_group TEXT NOT NULL,
                primary_label TEXT NOT NULL,
                secondary_label TEXT,
                share_primary REAL,
                share_secondary REAL,
                prime_share_primary REAL,
                prime_share_secondary REAL,
                price_per_sec_eur REAL NOT NULL,
                FOREIGN KEY(pricing_list_id) REFERENCES pricing_lists(id) ON DELETE CASCADE
            )""")

        # -------- Campaigns / Waves --------
        db.execute("""
        CREATE TABLE IF NOT EXISTS campaigns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            pricing_list_id INTEGER NOT NULL,
            start_date TEXT,
            end_date TEXT,
            status TEXT DEFAULT 'draft',
            FOREIGN KEY(pricing_list_id) REFERENCES pricing_lists(id) ON DELETE RESTRICT
        )""")
        
        # TVC (TV Commercials) table
        db.execute("""
        CREATE TABLE IF NOT EXISTS tvcs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            campaign_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            duration INTEGER NOT NULL, -- duration in seconds
            FOREIGN KEY(campaign_id) REFERENCES campaigns(id) ON DELETE CASCADE
        )""")
        db.execute("""
        CREATE TABLE IF NOT EXISTS waves (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            campaign_id INTEGER NOT NULL,
            name TEXT,
            start_date TEXT,
            end_date TEXT,
            FOREIGN KEY(campaign_id) REFERENCES campaigns(id) ON DELETE CASCADE
        )""")
        db.execute("""
        CREATE TABLE IF NOT EXISTS wave_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            wave_id INTEGER NOT NULL,
            owner TEXT NOT NULL,
            target_group TEXT NOT NULL,
            primary_label TEXT NOT NULL,
            secondary_label TEXT,
            share_primary REAL,
            share_secondary REAL,
            prime_share_primary REAL,
            prime_share_secondary REAL,
            price_per_sec_eur REAL NOT NULL,
            trps REAL NOT NULL,
            tvc_id INTEGER,
            FOREIGN KEY(wave_id) REFERENCES waves(id) ON DELETE CASCADE,
            FOREIGN KEY(tvc_id) REFERENCES tvcs(id) ON DELETE SET NULL
        )""")

        db.commit()
        
        # discount table
        
        db.execute("""
        CREATE TABLE IF NOT EXISTS discounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            campaign_id INTEGER,
            wave_id INTEGER,
            discount_type TEXT CHECK(discount_type IN ('client', 'agency')) NOT NULL,
            discount_percentage REAL NOT NULL,
            FOREIGN KEY(campaign_id) REFERENCES campaigns(id) ON DELETE CASCADE,
            FOREIGN KEY(wave_id) REFERENCES waves(id) ON DELETE CASCADE
        )
        """)
        db.commit()

def migrate_add_tvc_id_to_wave_items():
    """Add tvc_id column to wave_items table if it doesn't exist"""
    with get_db() as db:
        # Check if tvc_id column already exists
        cursor = db.execute("PRAGMA table_info(wave_items)")
        columns = [row[1] for row in cursor.fetchall()]
        
        if 'tvc_id' not in columns:
            # Add the tvc_id column
            db.execute("ALTER TABLE wave_items ADD COLUMN tvc_id INTEGER")
            db.commit()
            print("Added tvc_id column to wave_items table")

def migrate_add_campaign_fields():
    """Add missing fields to campaigns table"""
    with get_db() as db:
        cursor = db.execute("PRAGMA table_info(campaigns)")
        columns = [row[1] for row in cursor.fetchall()]
        
        new_columns = [
            ("agency", "TEXT"),
            ("client", "TEXT"),
            ("product", "TEXT"),
            ("country", "TEXT DEFAULT 'Lietuva'")
        ]
        
        for col_name, col_type in new_columns:
            if col_name not in columns:
                db.execute(f"ALTER TABLE campaigns ADD COLUMN {col_name} {col_type}")
                print(f"Added {col_name} column to campaigns table")
        
        db.commit()

def migrate_add_wave_item_fields():
    """Add all missing fields from Excel to wave_items table"""
    with get_db() as db:
        cursor = db.execute("PRAGMA table_info(wave_items)")
        columns = [row[1] for row in cursor.fetchall()]
        
        new_columns = [
            ("channel_id", "INTEGER"),
            ("channel_share", "REAL DEFAULT 0.75"),
            ("pt_zone_share", "REAL DEFAULT 0.55"),
            ("npt_zone_share", "REAL DEFAULT 0.45"),
            ("clip_duration", "INTEGER DEFAULT 10"),
            ("grp_planned", "REAL"),
            ("affinity1", "REAL"),
            ("affinity2", "REAL"),
            ("affinity3", "REAL"),
            ("gross_cpp_eur", "REAL"),
            ("duration_index", "REAL DEFAULT 1.0"),
            ("seasonal_index", "REAL DEFAULT 1.0"),
            ("trp_purchase_index", "REAL DEFAULT 0.95"),
            ("advance_purchase_index", "REAL DEFAULT 0.95"),
            ("web_index", "REAL DEFAULT 1.0"),
            ("advance_payment_index", "REAL DEFAULT 1.0"),
            ("loyalty_discount_index", "REAL DEFAULT 1.0"),
            ("position_index", "REAL DEFAULT 1.0"),
            ("gross_price_eur", "REAL"),
            ("client_discount", "REAL DEFAULT 0"),
            ("net_price_eur", "REAL"),
            ("agency_discount", "REAL DEFAULT 0"),
            ("net_net_price_eur", "REAL"),
            ("tg_size_thousands", "REAL DEFAULT 0"),
            ("tg_share_percent", "REAL DEFAULT 0"),
            ("tg_sample_size", "INTEGER DEFAULT 0"),
            ("daily_trp_distribution", "TEXT")  # JSON string for daily TRP values
        ]
        
        for col_name, col_type in new_columns:
            if col_name not in columns:
                db.execute(f"ALTER TABLE wave_items ADD COLUMN {col_name} {col_type}")
                print(f"Added {col_name} column to wave_items table")
        
        db.commit()

def migrate_add_pricing_indices():
    """Add duration and seasonal indices to pricing_list_items"""
    with get_db() as db:
        cursor = db.execute("PRAGMA table_info(pricing_list_items)")
        columns = [row[1] for row in cursor.fetchall()]
        
        new_columns = [
            ("duration_index", "REAL DEFAULT 1.0"),
            ("seasonal_index", "REAL DEFAULT 1.0"),
            ("tg_size_thousands", "REAL DEFAULT 0"),  # TG dydis tūkstančiais
            ("tg_share_percent", "REAL DEFAULT 0"),   # TG dalis procentais 
            ("tg_sample_size", "INTEGER DEFAULT 0")   # TG imties dydis
        ]
        
        for col_name, col_type in new_columns:
            if col_name not in columns:
                db.execute(f"ALTER TABLE pricing_list_items ADD COLUMN {col_name} {col_type}")
                print(f"Added {col_name} column to pricing_list_items table")
        
        db.commit()

# ---------------- Channel groups / channels ----------------

def upsert_channel_group(name: str) -> int:
    with get_db() as db:
        db.execute("INSERT OR IGNORE INTO channel_groups(name) VALUES (?)", (name,))
        row = db.execute("SELECT id FROM channel_groups WHERE name=?", (name,)).fetchone()
        return row["id"]

def list_channel_groups():
    with get_db() as db:
        rows = db.execute("SELECT id, name FROM channel_groups ORDER BY name").fetchall()
        return [dict(r) for r in rows]

def get_channel_group_by_id(group_id: int):
    with get_db() as db:
        row = db.execute("SELECT id, name FROM channel_groups WHERE id = ?", (group_id,)).fetchone()
        return dict(row) if row else None

def create_channel(channel_group_id: int, name: str, size: str):
    size = size.lower()
    if size not in ("big","small"):
        raise ValueError("size must be 'big' or 'small'")
    with get_db() as db:
        db.execute("""
            INSERT INTO channels(channel_group_id, name, size)
            VALUES (?,?,?)
        """, (channel_group_id, name, size))

def list_channels(channel_group_id: int | None = None):
    with get_db() as db:
        if channel_group_id:
            rows = db.execute("""
                SELECT c.id, c.name, c.size, c.channel_group_id, g.name AS group_name
                FROM channels c
                JOIN channel_groups g ON g.id=c.channel_group_id
                WHERE c.channel_group_id=?
                ORDER BY (size='big') DESC, c.name
            """, (channel_group_id,)).fetchall()
        else:
            rows = db.execute("""
                SELECT c.id, c.name, c.size, c.channel_group_id, g.name AS group_name
                FROM channels c
                JOIN channel_groups g ON g.id=c.channel_group_id
                ORDER BY g.name, (size='big') DESC, c.name
            """).fetchall()
        return [dict(r) for r in rows]

def list_all_channels():
    """Get all channels from all groups for campaign forms"""
    return list_channels()

def get_trp_rate_item(owner: str, target_group: str):
    """Get TRP rate item for specific owner and target group"""
    with get_db() as db:
        row = db.execute("""
            SELECT * FROM trp_rates 
            WHERE owner = ? AND target_group = ?
        """, (owner, target_group)).fetchone()
        if row:
            rate = dict(row)
            # Add default TG data since TRP rates don't have these fields
            rate["tg_size_thousands"] = 100.0  # Default TG size
            rate["tg_share_percent"] = 15.0    # Default TG share
            rate["tg_sample_size"] = 500       # Default sample size
            return rate
        return None

def update_channel(channel_id: int, *, name: str | None = None, size: str | None = None):
    sets, args = [], []
    if name is not None:
        sets.append("name=?"); args.append(name)
    if size is not None:
        size = size.lower()
        if size not in ("big","small"):
            raise ValueError("size must be 'big' or 'small'")
        sets.append("size=?"); args.append(size)
    if not sets:
        return
    args.append(channel_id)
    with get_db() as db:
        db.execute(f"UPDATE channels SET {', '.join(sets)} WHERE id=?", args)

def delete_channel(channel_id: int):
    with get_db() as db:
        db.execute("DELETE FROM channels WHERE id=?", (channel_id,))

def seed_channel_groups():
    amb_id = upsert_channel_group("AMB Baltics")
    mg_id  = upsert_channel_group("MG grupė")
    existing = { (r["channel_group_id"], r["name"]) for r in list_channels() }
    def ensure(gid, name, size):
        if (gid, name) not in existing:
            create_channel(gid, name, size)
    ensure(amb_id, "TV3", "big")
    ensure(amb_id, "TV6", "small")
    ensure(amb_id, "TV8", "small")
    ensure(amb_id, "TV3 Plus", "small")
    ensure(mg_id, "LNK", "big")

def update_channel_group(group_id: int, name: str):
    if not name or not name.strip():
        raise ValueError("name required")
    with get_db() as db:
        db.execute("UPDATE channel_groups SET name=? WHERE id=?", (name.strip(), group_id))

def delete_channel_group(group_id: int):
    with get_db() as db:
        db.execute("DELETE FROM channel_groups WHERE id=?", (group_id,))

# ---------------- TRP rates (legacy) ----------------

def _norm_number(x):
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return None
    s = s.replace("€", "").replace("%", "").replace(" ", "").replace(",", ".")
    return float(s)

def upsert_trp_rate(**k):
    if not k.get("owner") or not k.get("target_group") or not k.get("primary_label"):
        raise ValueError("owner, target_group, primary_label are required")
    if k.get("price_per_sec_eur") in (None, ""):
        raise ValueError("price_per_sec_eur is required")
    for f in ["share_primary","share_secondary","prime_share_primary","prime_share_secondary","price_per_sec_eur"]:
        if f in k:
            k[f] = _norm_number(k[f])
    with get_db() as db:
        db.execute("""
            INSERT INTO trp_rates (
                owner, target_group, primary_label, secondary_label,
                share_primary, share_secondary, prime_share_primary, prime_share_secondary,
                price_per_sec_eur
            ) VALUES (
                :owner, :target_group, :primary_label, :secondary_label,
                :share_primary, :share_secondary, :prime_share_primary, :prime_share_secondary,
                :price_per_sec_eur
            )
            ON CONFLICT(owner, target_group) DO UPDATE SET
                primary_label=excluded.primary_label,
                secondary_label=excluded.secondary_label,
                share_primary=excluded.share_primary,
                share_secondary=excluded.share_secondary,
                prime_share_primary=excluded.prime_share_primary,
                prime_share_secondary=excluded.prime_share_secondary,
                price_per_sec_eur=excluded.price_per_sec_eur
        """, k)

def list_trp_rates(owner=None):
    with get_db() as db:
        if owner:
            rows = db.execute("SELECT * FROM trp_rates WHERE owner=? ORDER BY target_group", (owner,)).fetchall()
        else:
            rows = db.execute("SELECT * FROM trp_rates ORDER BY owner, target_group").fetchall()
    return [dict(r) for r in rows]

def update_trp_rate_by_id(row_id, data: dict):
    if not data:
        return
    to_update = {}
    for key in ["owner","target_group","primary_label","secondary_label",
                "share_primary","share_secondary","prime_share_primary","prime_share_secondary",
                "price_per_sec_eur"]:
        if key in data:
            val = data[key]
            if key in {"share_primary","share_secondary","prime_share_primary","prime_share_secondary","price_per_sec_eur"}:
                val = _norm_number(val)
            to_update[key] = val
    if not to_update:
        return
    sets = ", ".join([f"{k}=?" for k in to_update.keys()])
    args = list(to_update.values()) + [row_id]
    with get_db() as db:
        db.execute(f"UPDATE trp_rates SET {sets} WHERE id=?", args)

def delete_trp_rate(row_id: int):
    with get_db() as db:
        db.execute("DELETE FROM trp_rates WHERE id=?", (row_id,))

# ---------------- Pricing lists (rate cards) ----------------

def create_pricing_list(name: str, auto_import: bool = True) -> int:
    """Create a new pricing list and optionally import TRP rates"""
    with get_db() as db:
        db.execute("INSERT INTO pricing_lists(name) VALUES (?)", (name,))
        list_id = db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        
        if auto_import:
            # Automatically import all TRP rates to the new pricing list
            # Get all TRP rates
            trp_rates = db.execute("SELECT * FROM trp_rates").fetchall()
            
            # Get channel groups mapping (name to ID)
            channel_groups = db.execute("SELECT id, name FROM channel_groups").fetchall()
            group_map = {g["name"]: g["id"] for g in channel_groups}
            
            for rate in trp_rates:
                # Convert owner name to group ID
                owner_id = group_map.get(rate["owner"], rate["owner"])
                
                # Import each TRP rate as pricing list item
                db.execute("""
                    INSERT OR REPLACE INTO pricing_list_items (
                        pricing_list_id, owner, target_group, 
                        primary_label, secondary_label,
                        share_primary, share_secondary, 
                        prime_share_primary, prime_share_secondary,
                        price_per_sec_eur
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    list_id,
                    str(owner_id),  # Store as string for compatibility
                    rate["target_group"],
                    rate["primary_label"],
                    rate["secondary_label"],
                    rate["share_primary"],
                    rate["share_secondary"],
                    rate["prime_share_primary"],
                    rate["prime_share_secondary"],
                    rate["price_per_sec_eur"]
                ))
        
        return list_id

def import_trp_rates_to_pricing_list(pricing_list_id: int):
    """Import all TRP rates into a pricing list"""
    with get_db() as db:
        # Get all TRP rates
        trp_rates = db.execute("SELECT * FROM trp_rates").fetchall()
        
        # Get channel groups mapping (name to ID)
        channel_groups = db.execute("SELECT id, name FROM channel_groups").fetchall()
        group_map = {g["name"]: g["id"] for g in channel_groups}
        
        for rate in trp_rates:
            # Convert owner name to group ID
            owner_id = group_map.get(rate["owner"], rate["owner"])
            
            # Import each TRP rate as pricing list item
            db.execute("""
                INSERT OR REPLACE INTO pricing_list_items (
                    pricing_list_id, owner, target_group, 
                    primary_label, secondary_label,
                    share_primary, share_secondary, 
                    prime_share_primary, prime_share_secondary,
                    price_per_sec_eur
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                pricing_list_id,
                str(owner_id),  # Store as string for compatibility
                rate["target_group"],
                rate["primary_label"],
                rate["secondary_label"],
                rate["share_primary"],
                rate["share_secondary"],
                rate["prime_share_primary"],
                rate["prime_share_secondary"],
                rate["price_per_sec_eur"]
            ))

def migrate_trp_rates_to_pricing_list(name: str) -> int:
    """Create a new pricing list and migrate all TRP rates to it"""
    return create_pricing_list(name, auto_import=True)

def list_pricing_lists():
    with get_db() as db:
        rows = db.execute("SELECT id, name FROM pricing_lists ORDER BY name").fetchall()
        return [dict(r) for r in rows]

def update_pricing_list(pl_id: int, name: str):
    with get_db() as db:
        db.execute("UPDATE pricing_lists SET name=? WHERE id=?", (name, pl_id))

def delete_pricing_list(pl_id: int):
    with get_db() as db:
        db.execute("DELETE FROM pricing_lists WHERE id=?", (pl_id,))

def upsert_pricing_list_item(**k):
    # expects: pricing_list_id, owner, target_group, primary_label, secondary_label,
    # shares, primes, price_per_sec_eur
    for f in ["share_primary","share_secondary","prime_share_primary","prime_share_secondary","price_per_sec_eur"]:
        if f in k:
            k[f] = _norm_number(k[f])
    with get_db() as db:
        db.execute("""
        INSERT INTO pricing_list_items (
            pricing_list_id, owner, target_group, primary_label, secondary_label,
            share_primary, share_secondary, prime_share_primary, prime_share_secondary, price_per_sec_eur
        ) VALUES (
            :pricing_list_id, :owner, :target_group, :primary_label, :secondary_label,
            :share_primary, :share_secondary, :prime_share_primary, :prime_share_secondary, :price_per_sec_eur
        )
        ON CONFLICT(pricing_list_id, owner, target_group) DO UPDATE SET
            primary_label=excluded.primary_label,
            secondary_label=excluded.secondary_label,
            share_primary=excluded.share_primary,
            share_secondary=excluded.share_secondary,
            prime_share_primary=excluded.prime_share_primary,
            prime_share_secondary=excluded.prime_share_secondary,
            price_per_sec_eur=excluded.price_per_sec_eur
        """, k)

def list_pricing_list_items(pl_id: int):
    with get_db() as db:
        rows = db.execute("""
            SELECT * FROM pricing_list_items
            WHERE pricing_list_id=?
            ORDER BY owner, target_group
        """, (pl_id,)).fetchall()
        return [dict(r) for r in rows]

def create_pricing_list_item(**k):
    # Create a new pricing list item (always INSERT, never UPDATE)
    for f in ["share_primary","share_secondary","prime_share_primary","prime_share_secondary","price_per_sec_eur"]:
        if f in k:
            k[f] = _norm_number(k[f])
    with get_db() as db:
        cursor = db.execute("""
        INSERT INTO pricing_list_items (
            pricing_list_id, owner, target_group, primary_label, secondary_label,
            share_primary, share_secondary, prime_share_primary, prime_share_secondary, price_per_sec_eur
        ) VALUES (
            :pricing_list_id, :owner, :target_group, :primary_label, :secondary_label,
            :share_primary, :share_secondary, :prime_share_primary, :prime_share_secondary, :price_per_sec_eur
        )
        """, k)
        return cursor.lastrowid

def delete_pricing_list_item(item_id: int):
    with get_db() as db:
        db.execute("DELETE FROM pricing_list_items WHERE id=?", (item_id,))

def update_pricing_list_item(item_id: int, data: dict):
    # Convert channel_group_id to owner if present
    if 'channel_group_id' in data:
        data['owner'] = str(data['channel_group_id'])
        del data['channel_group_id']
    
    # Normalize numeric fields
    for f in ["share_primary","share_secondary","prime_share_primary","prime_share_secondary","price_per_sec_eur"]:
        if f in data:
            data[f] = _norm_number(data[f])
    
    # Build UPDATE statement dynamically
    fields = []
    values = []
    for key, value in data.items():
        if key not in ['id', 'pricing_list_id']:  # Don't update these fields
            fields.append(f"{key}=?")
            values.append(value)
    
    if not fields:
        return  # Nothing to update
    
    values.append(item_id)  # Add item_id for WHERE clause
    
    with get_db() as db:
        db.execute(f"UPDATE pricing_list_items SET {','.join(fields)} WHERE id=?", values)

def get_pricing_item(pl_id: int, owner: str, target_group: str):
    with get_db() as db:
        row = db.execute("""
            SELECT * FROM pricing_list_items
            WHERE pricing_list_id=? AND owner=? AND target_group=?
        """, (pl_id, owner, target_group)).fetchone()
        return dict(row) if row else None

def list_pricing_owners(pl_id: int):
    with get_db() as db:
        rows = db.execute("""
            SELECT DISTINCT owner FROM pricing_list_items
            WHERE pricing_list_id=? ORDER BY owner
        """, (pl_id,)).fetchall()
        return [r["owner"] for r in rows]

def list_pricing_targets(pl_id: int, owner: str):
    with get_db() as db:
        rows = db.execute("""
            SELECT target_group FROM pricing_list_items
            WHERE pricing_list_id=? AND owner=? ORDER BY target_group
        """, (pl_id, owner)).fetchall()
        return [r["target_group"] for r in rows]

# ---------------- Campaigns / Waves ----------------

def create_campaign(name: str, start_date: str | None, end_date: str | None, 
                   agency: str = "", client: str = "", product: str = "", 
                   country: str = "Lietuva", status: str = "draft") -> int:
    with get_db() as db:
        db.execute("""
            INSERT INTO campaigns(name, start_date, end_date, 
                                agency, client, product, country, status)
            VALUES (?,?,?,?,?,?,?,?)
        """, (name, start_date, end_date, agency, client, product, country, status))
        return db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]

def list_campaigns():
    with get_db() as db:
        rows = db.execute("""
            SELECT c.*, NULL AS pricing_list_name
            FROM campaigns c
            ORDER BY c.id DESC
        """).fetchall()
        return [dict(r) for r in rows]

def update_campaign(cid: int, data: dict):
    sets, args = [], []
    for k in ["name","pricing_list_id","start_date","end_date","status"]:
        if k in data:
            sets.append(f"{k}=?"); args.append(data[k])
    if not sets:
        return
    args.append(cid)
    with get_db() as db:
        db.execute(f"UPDATE campaigns SET {', '.join(sets)} WHERE id=?", args)

def delete_campaign(cid: int):
    with get_db() as db:
        # Delete TRP distribution data first
        delete_trp_distribution(cid)
        # Delete campaign (waves and wave_items will cascade)
        db.execute("DELETE FROM campaigns WHERE id=?", (cid,))

def create_wave(campaign_id: int, name: str | None, start_date: str | None, end_date: str | None) -> int:
    with get_db() as db:
        db.execute("""
            INSERT INTO waves(campaign_id, name, start_date, end_date)
            VALUES (?,?,?,?)
        """, (campaign_id, name, start_date, end_date))
        return db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]

def list_waves(campaign_id: int):
    with get_db() as db:
        rows = db.execute("""
            SELECT * FROM waves WHERE campaign_id=? ORDER BY id
        """, (campaign_id,)).fetchall()
        return [dict(r) for r in rows]

def update_wave(wid: int, data: dict):
    sets, args = [], []
    for k in ["name","start_date","end_date"]:
        if k in data:
            sets.append(f"{k}=?"); args.append(data[k])
    if not sets:
        return
    args.append(wid)
    with get_db() as db:
        db.execute(f"UPDATE waves SET {', '.join(sets)} WHERE id=?", args)

def list_waves_for_deletion_sync(wid: int):
    with get_db() as db:
        row = db.execute("""
            SELECT w.name, w.campaign_id
            FROM waves w
            WHERE w.id=?
        """, (wid,)).fetchone()
        return dict(row) if row else None

def delete_wave(wid: int):
    with get_db() as db:
        db.execute("DELETE FROM waves WHERE id=?", (wid,))

def _pricing_list_id_for_wave(wave_id: int) -> int | None:
    with get_db() as db:
        row = db.execute("""
            SELECT c.pricing_list_id
            FROM waves w
            JOIN campaigns c ON c.id=w.campaign_id
            WHERE w.id=?
        """, (wave_id,)).fetchone()
        return row["pricing_list_id"] if row else None

def list_wave_items(wave_id: int):
    with get_db() as db:
        rows = db.execute("""
            SELECT * FROM wave_items WHERE wave_id=? ORDER BY id
        """, (wave_id,)).fetchall()
        return [dict(r) for r in rows]

def create_wave_item_prefill(wave_id: int, owner: str, target_group: str, trps: float, tvc_id: int = None) -> int:
    pl_id = _pricing_list_id_for_wave(wave_id)
    if not pl_id:
        raise ValueError("Pricing list not found for wave")
    rate = get_pricing_item(pl_id, owner, target_group)
    if not rate:
        raise ValueError("Rate not found in pricing list for given owner/target_group")
    
    # Validate TVC belongs to this wave's campaign if provided
    if tvc_id is not None:
        with get_db() as db:
            tvc_check = db.execute("""
            SELECT t.campaign_id, w.campaign_id as wave_campaign_id 
            FROM tvcs t, waves w 
            WHERE t.id = ? AND w.id = ?
            """, (tvc_id, wave_id)).fetchone()
            
            if not tvc_check or tvc_check['campaign_id'] != tvc_check['wave_campaign_id']:
                raise ValueError("TVC doesn't belong to this campaign")
    
    with get_db() as db:
        db.execute("""
            INSERT INTO wave_items(
                wave_id, owner, target_group, primary_label, secondary_label,
                share_primary, share_secondary, prime_share_primary, prime_share_secondary,
                price_per_sec_eur, trps, tvc_id
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            wave_id, owner, target_group, rate["primary_label"], rate["secondary_label"],
            rate["share_primary"], rate["share_secondary"], rate["prime_share_primary"], rate["prime_share_secondary"],
            rate["price_per_sec_eur"], _norm_number(trps), tvc_id
        ))
        return db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]

def create_wave_item_excel(wave_id: int, excel_data: dict) -> int:
    """Create a wave item with Excel-style data structure"""
        
    # Get pricing info from TRP rates based on channel_group (which is the owner) and target_group
    rate = get_trp_rate_item(excel_data["channel_group"], excel_data["target_group"])
    
    # Calculate derived values
    # GRP Planned = TRP × 100 / affinity1 (correct Excel formula)
    affinity1 = excel_data.get("affinity1")
    if affinity1 and affinity1 != 0:
        grp_planned = excel_data["trps"] * 100 / affinity1
    else:
        grp_planned = 0  # Cannot calculate without valid affinity1
    # CPP = price per second (from rate list, no clip duration multiplication)
    gross_cpp_eur = rate["price_per_sec_eur"] if rate else 1.0
    
    # Get indices from database if available, otherwise use form values
    # Get wave dates for seasonal index calculation
    with get_db() as db:
        wave_data = db.execute("SELECT start_date, end_date FROM waves WHERE id = ?", (wave_id,)).fetchone()
        wave_start_date = wave_data["start_date"] if wave_data else None
        wave_end_date = wave_data["end_date"] if wave_data else None
    
    # Get indices from database using channel group and wave date range
    db_indices = get_indices_for_wave_item(excel_data["channel_group"], excel_data["clip_duration"], wave_start_date, wave_end_date)
    
    # Use database indices if available, otherwise fall back to form values
    duration_index = db_indices.get("duration_index", excel_data.get("duration_index", 1.25))
    seasonal_index = db_indices.get("seasonal_index", excel_data.get("seasonal_index", 0.9))
    
    # Calculate gross price with all indices and clip duration
    gross_price_eur = (excel_data["trps"] * gross_cpp_eur * excel_data["clip_duration"] *
                      duration_index * seasonal_index * 
                      excel_data["trp_purchase_index"] * excel_data["advance_purchase_index"] * 
                      excel_data["position_index"])
    
    # Calculate net prices
    net_price_eur = gross_price_eur * (1 - excel_data["client_discount"] / 100)
    net_net_price_eur = net_price_eur * (1 - excel_data["agency_discount"] / 100)
    
    with get_db() as db:
        db.execute("""
            INSERT INTO wave_items(
                wave_id, target_group, trps, channel_id, channel_share, pt_zone_share, clip_duration, tvc_id,
                grp_planned, affinity1, affinity2, affinity3, gross_cpp_eur, duration_index,
                seasonal_index, trp_purchase_index, advance_purchase_index, position_index,
                gross_price_eur, client_discount, net_price_eur, agency_discount, net_net_price_eur,
                tg_size_thousands, tg_share_percent, tg_sample_size,
                owner, primary_label, secondary_label, share_primary, share_secondary,
                prime_share_primary, prime_share_secondary, price_per_sec_eur
            )
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            wave_id, excel_data["target_group"], _norm_number(excel_data["trps"]), None,  # channel_id set to None since we use channel_group
            excel_data["channel_share"], excel_data["pt_zone_share"], excel_data["clip_duration"], 
            excel_data.get("tvc_id"),  # TVC ID from form
            grp_planned, excel_data.get("affinity1"), excel_data.get("affinity2"), excel_data.get("affinity3"),
            gross_cpp_eur, duration_index, seasonal_index,
            excel_data["trp_purchase_index"], excel_data["advance_purchase_index"], excel_data["position_index"],
            gross_price_eur, excel_data["client_discount"], net_price_eur, 
            excel_data["agency_discount"], net_net_price_eur,
            # TG data from Excel/form, fallback to TRP rates, then defaults
            excel_data.get("tg_size_thousands") or (rate.get("tg_size_thousands", 0) if rate else 0),
            excel_data.get("tg_share_percent") or (rate.get("tg_share_percent", 0) if rate else 0), 
            excel_data.get("tg_sample_size") or (rate.get("tg_sample_size", 0) if rate else 0),
            # Use channel_group as owner
            excel_data["channel_group"], 
            rate["primary_label"] if rate else "N/A",
            rate["secondary_label"] if rate else None,
            rate["share_primary"] if rate else 0,
            rate["share_secondary"] if rate else 0,
            rate["prime_share_primary"] if rate else 0,
            rate["prime_share_secondary"] if rate else 0,
            rate["price_per_sec_eur"] if rate else gross_cpp_eur
        ))
        return db.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]

def update_wave_item(item_id: int, data: dict):
    print(f"DEBUG: update_wave_item called with item_id={item_id}, data={data}")
    # allow overriding any snapped values including discounts and Excel structure fields
    numeric = {"share_primary","share_secondary","prime_share_primary","prime_share_secondary","price_per_sec_eur","trps","client_discount","agency_discount",
               "channel_share","pt_zone_share","clip_duration","affinity1","affinity2","affinity3",
               "duration_index","seasonal_index","trp_purchase_index","advance_purchase_index","web_index","advance_payment_index","loyalty_discount_index","position_index"}
    sets, args = [], []
    for k in ["owner","target_group","primary_label","secondary_label",
              "share_primary","share_secondary","prime_share_primary","prime_share_secondary",
              "price_per_sec_eur","trps","client_discount","agency_discount",
              "channel_share","pt_zone_share","clip_duration","affinity1","affinity2","affinity3",
              "duration_index","seasonal_index","trp_purchase_index","advance_purchase_index","web_index","advance_payment_index","loyalty_discount_index","position_index"]:
        if k in data:
            v = _norm_number(data[k]) if k in numeric else data[k]
            sets.append(f"{k}=?"); args.append(v)
    
    # Recalculate prices and GRP if relevant fields were updated
    need_price_recalc = any(field in data for field in ["client_discount", "agency_discount", "trps", "clip_duration", "trp_purchase_index", "advance_purchase_index", "web_index", "advance_payment_index", "loyalty_discount_index", "position_index", "duration_index", "seasonal_index"])
    need_grp_recalc = any(field in data for field in ["trps", "affinity1"])
    
    if need_price_recalc or need_grp_recalc:
        with get_db() as db:
            # Get current item data
            item = db.execute("SELECT * FROM wave_items WHERE id = ?", (item_id,)).fetchone()
            if item:
                # Get updated values or use existing ones (SQLite Row uses [] not .get())
                trps = data.get("trps", item["trps"] or 0)
                gross_cpp = item["gross_cpp_eur"] or 0
                
                # Get updated indices or use existing ones
                duration_index = data.get("duration_index", item["duration_index"] or 1.0)
                seasonal_index = data.get("seasonal_index", item["seasonal_index"] or 1.0)
                trp_purchase_index = data.get("trp_purchase_index", item["trp_purchase_index"] or 0.95)
                advance_purchase_index = data.get("advance_purchase_index", item["advance_purchase_index"] or 0.95)
                web_index = data.get("web_index", item["web_index"] or 1.0)
                advance_payment_index = data.get("advance_payment_index", item["advance_payment_index"] or 1.0)
                loyalty_discount_index = data.get("loyalty_discount_index", item["loyalty_discount_index"] or 1.0)
                position_index = data.get("position_index", item["position_index"] or 1.0)
                
                # Recalculate gross price with all indices and clip duration
                clip_duration = data.get("clip_duration", item["clip_duration"] or 10)
                gross_price = (trps * gross_cpp * clip_duration * duration_index * seasonal_index *
                             trp_purchase_index * advance_purchase_index * web_index *
                             advance_payment_index * loyalty_discount_index * position_index)
                
                # Get discounts
                client_discount = data.get("client_discount", item["client_discount"] or 0)
                agency_discount = data.get("agency_discount", item["agency_discount"] or 0)
                
                # Calculate net prices
                net_price = gross_price * (1 - client_discount / 100)
                net_net_price = net_price * (1 - agency_discount / 100)
                
                # Recalculate GRP if needed
                if need_grp_recalc:
                    updated_trps = data.get("trps", item["trps"] or 0)
                    updated_affinity1 = data.get("affinity1", item["affinity1"])
                    
                    if updated_affinity1 and updated_affinity1 != 0:
                        grp_planned = updated_trps * 100 / updated_affinity1
                    else:
                        grp_planned = 0
                    
                    sets.append("grp_planned=?"); args.append(grp_planned)
                
                # Add price updates to sets
                sets.append("gross_price_eur=?"); args.append(gross_price)
                sets.append("net_price_eur=?"); args.append(net_price)
                sets.append("net_net_price_eur=?"); args.append(net_net_price)
    
    if not sets:
        print(f"DEBUG: No fields to update for item_id={item_id}")
        return
        
    args.append(item_id)
    sql = f"UPDATE wave_items SET {', '.join(sets)} WHERE id=?"
    print(f"DEBUG: Executing SQL: {sql} with args: {args}")
    
    with get_db() as db:
        db.execute(sql, args)
        print(f"DEBUG: Update completed for item_id={item_id}")

def delete_wave_item(item_id: int):
    with get_db() as db:
        db.execute("DELETE FROM wave_items WHERE id=?", (item_id,))

def recalculate_wave_item_prices_with_discounts(wave_id: int):
    """Recalculate all wave item prices using wave-level discounts"""
    with get_db() as db:
        # Get wave-level discounts
        discounts = db.execute("""
            SELECT discount_type, discount_percentage 
            FROM discounts 
            WHERE wave_id = ?
        """, (wave_id,)).fetchall()
        
        client_discount = 0
        agency_discount = 0
        
        for discount in discounts:
            if discount["discount_type"] == "client":
                client_discount = discount["discount_percentage"]
            elif discount["discount_type"] == "agency":
                agency_discount = discount["discount_percentage"]
        
        # Get all wave items for this wave
        items = db.execute("SELECT * FROM wave_items WHERE wave_id = ?", (wave_id,)).fetchall()
        
        for item in items:
            # Recalculate prices with wave-level discounts
            gross_price = item["gross_price_eur"] or 0
            net_price = gross_price * (1 - client_discount / 100)
            net_net_price = net_price * (1 - agency_discount / 100)
            
            # Update the item with new calculated prices and discounts
            db.execute("""
                UPDATE wave_items 
                SET client_discount = ?, agency_discount = ?, 
                    net_price_eur = ?, net_net_price_eur = ?
                WHERE id = ?
            """, (client_discount, agency_discount, net_price, net_net_price, item["id"]))
        
        db.commit()

# ---------------- TVCs (TV Commercials) ----------------

def create_tvc(campaign_id: int, name: str, duration: int):
    """Create a new TVC for a campaign"""
    if not name.strip():
        raise ValueError("TVC name is required")
    if duration <= 0:
        raise ValueError("Duration must be positive")
    
    with get_db() as db:
        cursor = db.execute("""
        INSERT INTO tvcs (campaign_id, name, duration)
        VALUES (?, ?, ?)
        """, (campaign_id, name.strip(), duration))
        return cursor.lastrowid

def list_campaign_tvcs(campaign_id: int):
    """Get all TVCs for a campaign"""
    with get_db() as db:
        rows = db.execute("""
        SELECT * FROM tvcs WHERE campaign_id = ? ORDER BY name
        """, (campaign_id,)).fetchall()
        return [dict(r) for r in rows]

def update_tvc(tvc_id: int, name: str = None, duration: int = None):
    """Update TVC name and/or duration"""
    updates = []
    values = []
    
    if name is not None:
        if not name.strip():
            raise ValueError("TVC name cannot be empty")
        updates.append("name = ?")
        values.append(name.strip())
    
    if duration is not None:
        if duration <= 0:
            raise ValueError("Duration must be positive")
        updates.append("duration = ?")
        values.append(duration)
    
    if not updates:
        return
    
    values.append(tvc_id)
    
    with get_db() as db:
        db.execute(f"UPDATE tvcs SET {', '.join(updates)} WHERE id = ?", values)

def delete_tvc(tvc_id: int):
    """Delete a TVC"""
    with get_db() as db:
        db.execute("DELETE FROM tvcs WHERE id = ?", (tvc_id,))

# ---------------- Discounts ----------------

def create_discount(campaign_id: int = None, wave_id: int = None, discount_type: str = "client", discount_percentage: float = 0.0):
    """Create a discount for a campaign or wave"""
    if not campaign_id and not wave_id:
        raise ValueError("Either campaign_id or wave_id must be provided")
    if discount_type not in ['client', 'agency']:
        raise ValueError("discount_type must be 'client' or 'agency'")
    
    with get_db() as db:
        cursor = db.execute("""
        INSERT INTO discounts (campaign_id, wave_id, discount_type, discount_percentage)
        VALUES (?, ?, ?, ?)
        """, (campaign_id, wave_id, discount_type, discount_percentage))
        return cursor.lastrowid

def get_discounts_for_campaign(campaign_id: int):
    """Get all discounts for a campaign (both campaign-level and wave-level)"""
    with get_db() as db:
        rows = db.execute("""
        SELECT d.*, w.name as wave_name 
        FROM discounts d
        LEFT JOIN waves w ON d.wave_id = w.id
        WHERE d.campaign_id = ? OR d.wave_id IN (
            SELECT id FROM waves WHERE campaign_id = ?
        )
        ORDER BY d.discount_type, d.wave_id
        """, (campaign_id, campaign_id)).fetchall()
        return [dict(r) for r in rows]

def get_discounts_for_wave(wave_id: int):
    """Get all discounts for a specific wave"""
    with get_db() as db:
        rows = db.execute("""
        SELECT * FROM discounts WHERE wave_id = ?
        ORDER BY discount_type
        """, (wave_id,)).fetchall()
        return [dict(r) for r in rows]

def update_discount(discount_id: int, discount_percentage: float):
    """Update discount percentage"""
    with get_db() as db:
        db.execute("UPDATE discounts SET discount_percentage = ? WHERE id = ?", 
                  (discount_percentage, discount_id))

def delete_discount(discount_id: int):
    """Delete a discount"""
    with get_db() as db:
        db.execute("DELETE FROM discounts WHERE id = ?", (discount_id,))

def calculate_wave_total_with_discounts(wave_id: int):
    """Calculate wave total cost with discounts applied"""
    with get_db() as db:
        # Get base cost from wave items
        items = db.execute("""
        SELECT price_per_sec_eur, trps FROM wave_items WHERE wave_id = ?
        """, (wave_id,)).fetchall()
        
        base_cost = sum(item['price_per_sec_eur'] * item['trps'] for item in items)
        
        # Get discounts for this wave
        discounts = get_discounts_for_wave(wave_id)
        
        client_discount = 0
        agency_discount = 0
        
        for discount in discounts:
            if discount['discount_type'] == 'client':
                client_discount = max(client_discount, discount['discount_percentage'])
            elif discount['discount_type'] == 'agency':
                agency_discount = max(agency_discount, discount['discount_percentage'])
        
        # Apply discounts sequentially
        client_cost = base_cost * (1 - client_discount / 100)
        agency_cost = client_cost * (1 - agency_discount / 100)
        
        return {
            'base_cost': base_cost,
            'client_cost': client_cost,
            'agency_cost': agency_cost,
            'client_discount_percent': client_discount,
            'agency_discount_percent': agency_discount
        }

# ---------------- Campaign Status ----------------

def update_campaign_status(campaign_id: int, status: str):
    """Update campaign status"""
    valid_statuses = ['draft', 'confirmed', 'orders_sent', 'active', 'completed']
    if status not in valid_statuses:
        raise ValueError(f"Status must be one of: {', '.join(valid_statuses)}")
    
    with get_db() as db:
        db.execute("UPDATE campaigns SET status = ? WHERE id = ?", (status, campaign_id))

def get_campaign_status(campaign_id: int):
    """Get campaign status"""
    with get_db() as db:
        row = db.execute("SELECT status FROM campaigns WHERE id = ?", (campaign_id,)).fetchone()
        return row['status'] if row else None

# ---------------- Report Generation ----------------

def get_campaign_report_data(campaign_id: int):
    """Get all data needed for campaign reports"""
    with get_db() as db:
        # Get campaign info (pricing_list_id might be NULL, so use LEFT JOIN)
        campaign = db.execute("""
        SELECT c.*, COALESCE(pl.name, 'Default') as pricing_list_name 
        FROM campaigns c
        LEFT JOIN pricing_lists pl ON c.pricing_list_id = pl.id
        WHERE c.id = ?
        """, (campaign_id,)).fetchone()
        
        if not campaign:
            return None
        
        # Get waves with items and costs
        waves_data = []
        waves = db.execute("SELECT * FROM waves WHERE campaign_id = ? ORDER BY start_date, name", (campaign_id,)).fetchall()
        
        for wave in waves:
            wave_dict = dict(wave)
            
            # Get wave items with TVC info
            items = db.execute("""
            SELECT wi.*, t.name as tvc_name, t.duration as tvc_duration
            FROM wave_items wi
            LEFT JOIN tvcs t ON wi.tvc_id = t.id
            WHERE wi.wave_id = ? 
            ORDER BY wi.owner, wi.target_group
            """, (wave['id'],)).fetchall()
            wave_dict['items'] = [dict(item) for item in items]
            
            # Get wave costs with discounts
            wave_dict['costs'] = calculate_wave_total_with_discounts(wave['id'])
            
            # Get discounts
            discounts = get_discounts_for_wave(wave['id'])
            wave_dict['discounts'] = discounts
            
            waves_data.append(wave_dict)
        
        return {
            'campaign': dict(campaign),
            'waves': waves_data
        }

# ---------------- TRP Calendar ----------------

def save_trp_distribution(campaign_id: int, trp_data: dict):
    """Save TRP distribution for a campaign"""
    with get_db() as db:
        for date_str, trp_value in trp_data.items():
            db.execute("""
                INSERT INTO trp_distribution (campaign_id, date, trp_value, updated_at)
                VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(campaign_id, date) DO UPDATE SET
                    trp_value = excluded.trp_value,
                    updated_at = CURRENT_TIMESTAMP
            """, (campaign_id, date_str, float(trp_value) if trp_value else 0.0))

def load_trp_distribution(campaign_id: int):
    """Load TRP distribution for a campaign"""
    with get_db() as db:
        rows = db.execute("""
            SELECT date, trp_value FROM trp_distribution 
            WHERE campaign_id = ? AND trp_value > 0
            ORDER BY date
        """, (campaign_id,)).fetchall()
        
        return {row['date']: row['trp_value'] for row in rows}

def delete_trp_distribution(campaign_id: int):
    """Delete all TRP distribution data for a campaign"""
    with get_db() as db:
        db.execute("DELETE FROM trp_distribution WHERE campaign_id = ?", (campaign_id,))

# openpyxl imports moved inside export_channel_group_excel function
from io import BytesIO
import csv
import io

def generate_client_excel_report(campaign_id: int):
    """Generate Excel report for client (with client discounts applied)"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from datetime import datetime, timedelta
    import sys

    data = get_campaign_report_data(campaign_id)
    if not data:
        return None

    campaign = data['campaign']
    waves = data['waves']

    # Load TRP calendar data
    trp_data = load_trp_distribution(campaign_id)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TV Planas"
    
    # Professional styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Dark blue
    header_font = Font(color="FFFFFF", bold=True, size=11)
    wave_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")  # Light blue
    wave_font = Font(bold=True, color="1F4E79", size=10)
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow for totals
    border = Border(left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'), 
                   top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'))
    thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), 
                         top=Side(style='medium'), bottom=Side(style='medium'))
    
    current_row = 1
    
    # Campaign header
    ws.merge_cells(f'A{current_row}:P{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = f"TV KOMUNIKACIJOS PLANAS"
    cell.font = Font(size=18, bold=True, color="1F4E79")
    cell.alignment = Alignment(horizontal='center')
    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    current_row += 1

    # Campaign name
    ws.merge_cells(f'A{current_row}:P{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = f"KAMPANIJA: {campaign['name'].upper()}"
    cell.font = Font(size=14, bold=True, color="1F4E79")
    cell.alignment = Alignment(horizontal='center')
    current_row += 3
    
    # Campaign info section
    info_font = Font(bold=True, size=10, color="1F4E79")
    value_font = Font(size=10)
    
    # Create a bordered info section
    info_cells = [
        ("Kampanijos laikotarpis:", f"{campaign.get('start_date', '')} - {campaign.get('end_date', '')}"),
        ("Klientas:", campaign.get('client', '')),
        ("Agentūra:", campaign.get('agency', '')),
        ("Produktas:", campaign.get('product', '')),
        ("Kainoraštis:", campaign.get('pricing_list_name', '')),
        ("Statusas:", campaign.get('status', 'draft').replace('_', ' ').title())
    ]
    
    for label, value in info_cells:
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = info_font
        ws[f'B{current_row}'] = value
        ws[f'B{current_row}'].font = value_font
        current_row += 1
    
    # Main table starts at row 15 (after campaign info)
    current_row = 15
    
    # Table headers - keeping only the columns you want (16 columns total)
    headers = [
        'Pradžia', 'Pabaiga', 'Kanalų grupė', 'Perkama TG', 'TVC', 'Trukmė',
        'Kanalo dalis', 'PT zonos dalis', 'nPT zonos dalis', 'GRP plan.',
        'Gross CPP', 'Trukmės koeficientas', 'Sezoninis koeficientas',
        'Gross kaina', 'Kliento nuolaida %', 'Net kaina'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Set header row height to make it taller (CLIENT EXCEL EXPORT)
    ws.row_dimensions[current_row].height = 120

    current_row += 1
    
    # Data rows
    total_cost = 0
    row_count = 0
    light_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")  # Zebra striping
    
    for wave in waves:
        wave_start_row = current_row
        
        for item in wave['items']:
            item_cost = item['trps'] * item['price_per_sec_eur']
            
            # Calculate values
            tg_size = item.get('tg_size_thousands', 0)
            tg_share = item.get('tg_share_percent', 0)
            tg_sample = item.get('tg_sample_size', 0)
            channel_share = item.get('channel_share', 0.75)
            if channel_share < 1:  # If it's a decimal (0.75), convert to percentage
                channel_share = channel_share * 100
            pt_zone_share = item.get('pt_zone_share', 0.55)
            if pt_zone_share < 1:  # If it's a decimal (0.55), convert to percentage
                pt_zone_share = pt_zone_share * 100
            affinity1 = item.get('affinity1', 0)
            grp_planned = (item['trps'] * 100 / affinity1) if affinity1 > 0 else 0
            gross_cpp = item.get('gross_cpp_eur', item.get('price_per_sec_no_discount', item['price_per_sec_eur']))
            
            # Indices
            duration_idx = item.get('duration_index', 1.0)
            seasonal_idx = item.get('seasonal_index', 1.0)
            trp_purchase_idx = item.get('trp_purchase_index', 1.0)
            advance_idx = item.get('advance_purchase_index', 1.0)
            position_idx = item.get('position_index', 1.0)
            
            # Calculate base gross price (TRP * CPP * Duration) without multipliers
            # The Excel shows individual coefficients as separate rows, so we want the base calculation
            clip_duration = item.get('tvc_duration', item.get('clip_duration', 0))
            gross_price = item['trps'] * gross_cpp * clip_duration

            # Get actual client discount from item data
            client_discount = item.get('client_discount', 0)
            print(f"DEBUG EXCEL: TRP={item['trps']}, CPP={gross_cpp}, Duration={clip_duration}, Gross={gross_price}, Client_discount={client_discount}", file=sys.stderr, flush=True)
            net_price = gross_price * (1 - client_discount / 100)
            agency_discount = item.get('agency_discount', 0)
            net_net_price = net_price * (1 - agency_discount / 100)
            
            # Populate only the 16 remaining columns
            ws.cell(row=current_row, column=1).value = wave.get('start_date', '')  # Pradžia
            ws.cell(row=current_row, column=2).value = wave.get('end_date', '')  # Pabaiga
            ws.cell(row=current_row, column=3).value = item['owner']  # Kanalų grupė
            ws.cell(row=current_row, column=4).value = item['target_group']  # Perkama TG
            ws.cell(row=current_row, column=5).value = item.get('tvc_name', '-')  # TVC
            ws.cell(row=current_row, column=6).value = item.get('tvc_duration', item.get('clip_duration', 0))  # Trukmė
            ws.cell(row=current_row, column=7).value = channel_share / 100  # Kanalo dalis - as decimal
            ws.cell(row=current_row, column=8).value = pt_zone_share / 100  # PT zonos dalis - as decimal
            ws.cell(row=current_row, column=9).value = 0.45  # nPT zonos dalis - default value
            ws.cell(row=current_row, column=10).value = round(grp_planned, 2) if grp_planned > 0 else ""  # GRP plan.
            ws.cell(row=current_row, column=11).value = gross_cpp  # Gross CPP - as number
            ws.cell(row=current_row, column=12).value = duration_idx  # Trukmės koeficientas
            ws.cell(row=current_row, column=13).value = seasonal_idx  # Sezoninis koeficientas
            ws.cell(row=current_row, column=14).value = gross_price  # Gross kaina - as number
            ws.cell(row=current_row, column=15).value = client_discount  # Kl. nuol. % - show the actual percentage value
            ws.cell(row=current_row, column=16).value = net_price  # Net kaina - as number
            
            # Apply borders and formatting to all 16 columns
            for col in range(1, 17):
                cell = ws.cell(row=current_row, column=col)
                cell.border = border

                # Add center alignment for Pradžia, Pabaiga, and Kanalų grupė columns with text wrapping
                if col in [1, 2, 3]:  # Pradžia, Pabaiga, Kanalų grupę
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    # All other columns get text wrapping with left alignment
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                # Format numbers based on new 16-column structure
                if col in [7, 8, 9]:  # Percentage columns: Kanalo dalis, PT zonos dalis, nPT zonos dalis
                    cell.number_format = '0.00%'
                elif col in [12, 13]:  # Index columns: Trukmės koeficientas, Sezoninis koeficientas
                    cell.number_format = '0.00'
                elif col in [10, 11, 14, 16]:  # Number/Currency columns: GRP plan., Gross CPP, Gross kaina, Net kaina
                    cell.number_format = '#,##0.00'
                elif col == 15:  # Discount percentage column: Kl. nuol. %
                    cell.number_format = '0.0"%"'

            # Set row height to accommodate wrapped text
            ws.row_dimensions[current_row].height = 40

            current_row += 1
            row_count += 1
        
        # Wave summary with client discount
        if wave['items']:
            costs = wave['costs']
            client_cost = costs['client_cost']
            total_cost += client_cost
            
            ws.merge_cells(f'A{current_row}:W{current_row}')
            cell = ws[f'A{current_row}']
            discount_text = f" (-{costs['client_discount_percent']}%)" if costs['client_discount_percent'] > 0 else ""
            cell.value = f"Bangos suma{discount_text}:"
            cell.font = wave_font
            cell.fill = wave_fill
            cell.alignment = Alignment(horizontal='right')
            
            ws.cell(row=current_row, column=24).value = f"€{client_cost:.2f}"
            ws.cell(row=current_row, column=24).font = wave_font
            ws.cell(row=current_row, column=24).fill = wave_fill
            
            for col in range(1, 25):
                cell = ws.cell(row=current_row, column=col)
                cell.border = border
                cell.fill = wave_fill
            
            current_row += 2
    
    # Grand total - updated to span 16 columns with proper borders
    ws.merge_cells(f'A{current_row}:O{current_row}')  # Merge A to O (15 columns)
    cell = ws[f'A{current_row}']
    cell.value = "BENDRA KAMPANIJOS SUMA:"
    cell.font = Font(bold=True, size=12, color="1F4E79")
    cell.alignment = Alignment(horizontal='right')
    cell.fill = total_fill
    cell.border = thick_border

    # Total amount in the last column (P - column 16)
    total_cell = ws.cell(row=current_row, column=16)
    total_cell.value = f"€{total_cost:.2f}"
    total_cell.font = Font(bold=True, size=12, color="1F4E79")
    total_cell.fill = total_fill
    total_cell.border = thick_border

    # Add borders to all cells in the total row to ensure complete border coverage
    for col in range(1, 17):  # All 16 columns
        cell = ws.cell(row=current_row, column=col)
        cell.fill = total_fill
        cell.border = thick_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Set row height for the total row
    ws.row_dimensions[current_row].height = 45
    
    # Add Enhanced TRP Calendar at the same level as main table (to the right)
    if trp_data or (campaign.get('start_date') and campaign.get('end_date')):
        # Calendar will be positioned at the same level as main table (row 15+)
        cal_start_row = 15  # Same as main table start
        
        import json
        from datetime import datetime, timedelta
        if isinstance(trp_data, str):
            trp_data = json.loads(trp_data)
        
        # Get date range from campaign or extend to include all waves
        start_date_str = campaign.get('start_date', '')
        end_date_str = campaign.get('end_date', '')
        
        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
                
                # Check if any waves extend beyond campaign dates
                for wave in waves:
                    if wave.get('start_date'):
                        wave_start = datetime.strptime(wave['start_date'], '%Y-%m-%d')
                        if wave_start < start_date:
                            start_date = wave_start
                    if wave.get('end_date'):
                        wave_end = datetime.strptime(wave['end_date'], '%Y-%m-%d')
                        if wave_end > end_date:
                            end_date = wave_end
                
                # Show all active days - no date limiting, just make columns narrower
                calendar_limited = False  # We're not limiting dates anymore
                
                # Create horizontal calendar like in the UI - start from column AA (27) to not interfere with main table
                
                
                # Month headers row
                current_date = start_date
                col_idx = 35  # Start calendar from column AI - push to the right
                months = []
                month_spans = {}
                
                # Calculate month spans
                while current_date <= end_date:
                    month_key = current_date.strftime('%Y-%m')
                    month_name = current_date.strftime('%B %Y')
                    if month_key not in month_spans:
                        month_spans[month_key] = {'name': month_name, 'start_col': col_idx, 'days': 0}
                        months.append(month_key)
                    month_spans[month_key]['days'] += 1
                    current_date += timedelta(days=1)
                    col_idx += 1
                
                # Render month headers
                for month_key in months:
                    month_info = month_spans[month_key]
                    start_col = month_info['start_col']
                    end_col = start_col + month_info['days'] - 1
                    
                    # Set value first, then merge
                    ws.cell(row=cal_start_row, column=start_col).value = month_info['name']
                    
                    # Style the main cell before merging
                    cell = ws.cell(row=cal_start_row, column=start_col)
                    cell.font = Font(bold=True, size=10, color="1F4E79")  # Readable font
                    cell.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border
                    
                    # Merge after setting value and style (if needed)
                    if start_col != end_col:
                        ws.merge_cells(f'{openpyxl.utils.get_column_letter(start_col)}{cal_start_row}:{openpyxl.utils.get_column_letter(end_col)}{cal_start_row}')
                
                # Day numbers row
                current_date = start_date
                col_idx = 35  # Start from column AI
                while current_date <= end_date:
                    day_cell = ws.cell(row=cal_start_row + 1, column=col_idx)
                    day_cell.value = current_date.day
                    day_cell.font = Font(bold=True, size=10)  # Readable day numbers
                    day_cell.alignment = Alignment(horizontal='center')
                    
                    # Weekend styling
                    if current_date.weekday() in [5, 6]:  # Saturday, Sunday
                        day_cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                    else:
                        day_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    day_cell.border = border
                    
                    current_date += timedelta(days=1)
                    col_idx += 1
                
                # Week days row
                current_date = start_date
                col_idx = 35  # Start from column AI
                weekday_names = ['Pr', 'An', 'Tr', 'Kt', 'Pn', 'Št', 'Sk']
                while current_date <= end_date:
                    weekday_cell = ws.cell(row=cal_start_row + 2, column=col_idx)
                    weekday_cell.value = weekday_names[current_date.weekday()]
                    weekday_cell.font = Font(size=9)  # Readable weekday names
                    weekday_cell.alignment = Alignment(horizontal='center')
                    
                    # Weekend styling
                    if current_date.weekday() in [5, 6]:  # Saturday, Sunday
                        weekday_cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                        weekday_cell.font = Font(size=9, color="999999")  # Readable weekend font
                    else:
                        weekday_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    weekday_cell.border = border
                    
                    current_date += timedelta(days=1)
                    col_idx += 1
                
                # Wave rows with individual TRP distribution (showing TRP values per wave per day)
                row_idx = cal_start_row + 3
                valid_wave_count = 0
                for wave_idx, wave in enumerate(waves):
                    if wave.get('start_date') and wave.get('end_date'):
                        try:
                            wave_start = datetime.strptime(wave['start_date'], '%Y-%m-%d')
                            wave_end = datetime.strptime(wave['end_date'], '%Y-%m-%d')
                            
                            # Calculate total TRP for this wave
                            wave_total_trp = sum(item['trps'] for item in wave['items'] if item.get('trps', 0) > 0)
                            wave_days = (wave_end - wave_start).days + 1
                            daily_trp = wave_total_trp / wave_days if wave_days > 0 else 0
                            
                            current_date = start_date
                            col_idx = 35  # Start from column AI
                            
                            while current_date <= end_date:
                                wave_cell = ws.cell(row=row_idx, column=col_idx)
                                
                                if wave_start <= current_date <= wave_end:
                                    # Show TRP value for active days
                                    wave_cell.value = f"{daily_trp:.2f}" if daily_trp > 0 else ""
                                    wave_cell.font = Font(size=9, bold=True, color="FFFFFF")
                                    if current_date.weekday() in [5, 6]:  # Weekend
                                        wave_cell.fill = PatternFill(start_color="81C784", end_color="81C784", fill_type="solid")  # Light green
                                    else:
                                        wave_cell.fill = PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid")  # Green
                                    wave_cell.alignment = Alignment(horizontal='center')
                                else:
                                    wave_cell.value = ""
                                    if current_date.weekday() in [5, 6]:  # Weekend
                                        wave_cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
                                    else:
                                        wave_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                                
                                wave_cell.border = border
                                current_date += timedelta(days=1)
                                col_idx += 1
                            
                            row_idx += 1
                            valid_wave_count += 1
                        except ValueError as e:
                            # Skip waves with invalid dates
                            continue
                
                # Skip the combined TRP row - we now show individual TRP per wave
                trp_row = row_idx  # row_idx will be at the end of all wave rows
                
                # Column labels for the calendar
                ws.cell(row=cal_start_row + 2, column=col_idx + 1).value = "Savaitės dienos"
                ws.cell(row=cal_start_row + 2, column=col_idx + 1).font = Font(size=8, italic=True)
                
                # Wave labels with TRP totals
                label_row = cal_start_row + 3
                for wave_idx, wave in enumerate(waves):
                    if wave.get('start_date') and wave.get('end_date'):
                        try:
                            datetime.strptime(wave['start_date'], '%Y-%m-%d')
                            datetime.strptime(wave['end_date'], '%Y-%m-%d')
                            wave_total_trp = sum(item['trps'] for item in wave['items'] if item.get('trps', 0) > 0)
                            # Get channel group from first item in wave
                            channel_group = wave['items'][0]['owner'] if wave['items'] and wave['items'][0].get('owner') else f"Banga {wave_idx + 1}"
                            label = f"{channel_group} (TRP: {wave_total_trp:.2f})"
                            ws.cell(row=label_row, column=col_idx + 1).value = label
                            ws.cell(row=label_row, column=col_idx + 1).font = Font(size=9, color="1F4E79", bold=True)
                            label_row += 1
                        except ValueError:
                            continue
                
                # Total TRP summary calculated from all waves
                
                # Set row heights for calendar section to standard readable height
                for row in range(cal_start_row, trp_row + 3):  # All calendar rows
                    try:
                        ws.row_dimensions[row].height = 13  # Standard readable height
                    except:
                        pass  # Skip if row setting fails
                
                # Make calendar columns wider to show TRP values properly
                calendar_start_col = 35  # AI column - push calendar further to the right
                calendar_end_col = min(col_idx + 1, 60)  # Limit to reasonable range
                for col in range(calendar_start_col, calendar_end_col):
                    try:
                        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 4.5  # Wider for TRP values
                    except:
                        pass
                
            except Exception as e:
                # Fallback to simple table if date parsing fails
                print(f"Calendar generation error: {e}")  # For debugging
                try:
                    ws.cell(row=cal_start_row, column=1).value = "Data"
                    ws.cell(row=cal_start_row, column=2).value = "TRP"
                    for col in [1, 2]:
                        cell = ws.cell(row=cal_start_row, column=col)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = border
                    cal_start_row += 1
                    
                    if trp_data:
                        for date_str, trp_value in sorted(trp_data.items()):
                            try:
                                ws.cell(row=cal_start_row, column=1).value = date_str
                                ws.cell(row=cal_start_row, column=2).value = float(trp_value)
                                # Add borders for consistency
                                for col in [1, 2]:
                                    ws.cell(row=cal_start_row, column=col).border = border
                                cal_start_row += 1
                            except:
                                continue  # Skip problematic entries
                except Exception as fallback_error:
                    print(f"Fallback table creation failed: {fallback_error}")
                    # If even fallback fails, just add a simple message
                    try:
                        ws.cell(row=cal_start_row, column=1).value = "TRP data available - see campaign details"
                    except:
                        pass  # Give up gracefully
    
    
    # Set wider column widths for the 16 remaining columns to properly fit content
    ws.column_dimensions['A'].width = 16   # Pradžia
    ws.column_dimensions['B'].width = 16   # Pabaiga
    ws.column_dimensions['C'].width = 20   # Kanalų grupę
    ws.column_dimensions['D'].width = 18   # Perkama TG
    ws.column_dimensions['E'].width = 15   # TVC
    ws.column_dimensions['F'].width = 12   # Trukmé
    ws.column_dimensions['G'].width = 15   # Kanalo dalis
    ws.column_dimensions['H'].width = 15   # PT zonos dalis
    ws.column_dimensions['I'].width = 16   # nPT zonos dalis
    ws.column_dimensions['J'].width = 14   # GRP plan.
    ws.column_dimensions['K'].width = 14   # Gross CPP
    ws.column_dimensions['L'].width = 18   # Trukmés koeficientas
    ws.column_dimensions['M'].width = 18   # Sezoninis koeficientas
    ws.column_dimensions['N'].width = 16   # Gross kaina
    ws.column_dimensions['O'].width = 15   # Kliento nuolaida %
    ws.column_dimensions['P'].width = 16   # Net kaina
    
    # Save to BytesIO
    print(f"DEBUG: Saving workbook to BytesIO", file=sys.stderr, flush=True)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    print(f"DEBUG: Excel generation complete, returning buffer", file=sys.stderr, flush=True)
    return output

def generate_agency_csv_order(campaign_id: int):
    """Generate CSV order file for agency (with both client and agency discounts)"""
    data = get_campaign_report_data(campaign_id)
    if not data:
        return None
    
    campaign = data['campaign']
    waves = data['waves']
    
    # Load TRP calendar data
    trp_data = load_trp_distribution(campaign_id)
    
    csv_content = []
    csv_content.append(['# TV UŽSAKYMAS'])
    csv_content.append(['Kampanija', campaign['name']])
    csv_content.append(['Laikotarpis', f"{campaign.get('start_date', '')} - {campaign.get('end_date', '')}"])
    csv_content.append(['Kainoraštis', campaign.get('pricing_list_name', '')])
    csv_content.append(['Statusas', campaign.get('status', 'draft')])
    csv_content.append([])  # Empty row
    
    # Headers
    csv_content.append([
        'Banga', 'Laikotarpis', 'TVC', 'Trukmė', 'Savininkas', 'Tikslinė grupė', 'Kanalas', 
        'TRP', 'CPP €', 'TG dydis', 'Bazinė kaina €/sek', 'Kliento kaina €/sek', 'Agentūros kaina €/sek', 
        'Kliento nuolaida %', 'Agentūros nuolaida %', 'Galutinė suma €'
    ])
    
    total_agency_cost = 0
    
    for wave in waves:
        costs = wave['costs']
        
        for item in wave['items']:
            base_price = item['price_per_sec_eur']
            trps = item['trps']
            
            # Calculate discounted prices per second
            client_discount_percent = costs['client_discount_percent']
            agency_discount_percent = costs['agency_discount_percent']
            
            client_price_per_sec = base_price * (1 - client_discount_percent / 100)
            agency_price_per_sec = client_price_per_sec * (1 - agency_discount_percent / 100)
            
            final_cost = agency_price_per_sec * trps
            total_agency_cost += final_cost
            
            csv_content.append([
                wave['name'] or f"Banga {wave['id']}",
                f"{wave.get('start_date', '')} - {wave.get('end_date', '')}",
                item.get('tvc_name', '-'),
                f"{item.get('tvc_duration', item.get('clip_duration', 0))}s",
                item['owner'],
                item['target_group'],
                f"{item['primary_label']}{' + ' + item['secondary_label'] if item['secondary_label'] else ''}",
                trps,
                round(item.get('price_per_sec_no_discount', base_price), 2),
                f"{item.get('tg_size_thousands', 0)}k",
                round(base_price, 4),
                round(client_price_per_sec, 4),
                round(agency_price_per_sec, 4),
                client_discount_percent,
                agency_discount_percent,
                round(final_cost, 2)
            ])
    
    # Total
    csv_content.append([])
    csv_content.append(['BENDRA AGENTŪROS SUMA €', '', '', '', '', '', '', '', '', '', '', '', '', '', '', round(total_agency_cost, 2)])
    
    # Add TRP Calendar if exists
    if trp_data:
        csv_content.append([])
        csv_content.append(['# TRP KALENDORIUS'])
        csv_content.append(['Data', 'TRP'])
        
        import json
        if isinstance(trp_data, str):
            trp_data = json.loads(trp_data)
            
        for date_str, trp_value in sorted(trp_data.items()):
            csv_content.append([date_str, float(trp_value)])
    
    # Write CSV
    import csv as csv_module
    from io import StringIO
    
    # First write to string
    string_buffer = StringIO()
    writer = csv_module.writer(string_buffer, delimiter=';', lineterminator='\n')
    
    for row in csv_content:
        writer.writerow(row)
    
    # Convert to bytes with UTF-8 BOM
    csv_string = string_buffer.getvalue()
    output = BytesIO()
    output.write('\ufeff'.encode('utf-8'))  # BOM
    output.write(csv_string.encode('utf-8'))
    output.seek(0)
    
    return output

# ---------- INDICES MANAGEMENT ----------

def migrate_add_indices_tables():
    """Create tables for duration and seasonal indices management - LEGACY FUNCTION
    
    This function is now deprecated. Indices have been migrated to use channel_group_id
    instead of target_group. The new structure is already created by the migration script.
    """
    # This function is now a no-op since the migration has already been run
    # and the new channel_group-based structure is in place
    print("migrate_add_indices_tables: Indices tables already migrated to channel group structure")
    pass

def list_duration_indices():
    """Get all duration indices grouped by channel group"""
    with get_db() as db:
        return [dict(row) for row in db.execute(
            """SELECT di.*, cg.name as channel_group_name 
               FROM duration_indices di 
               JOIN channel_groups cg ON di.channel_group_id = cg.id 
               ORDER BY cg.name, di.duration_seconds"""
        ).fetchall()]

def list_seasonal_indices():
    """Get all seasonal indices grouped by channel group"""
    with get_db() as db:
        return [dict(row) for row in db.execute(
            """SELECT si.*, cg.name as channel_group_name 
               FROM seasonal_indices si 
               JOIN channel_groups cg ON si.channel_group_id = cg.id 
               ORDER BY cg.name, si.month"""
        ).fetchall()]

def list_position_indices():
    """Get all position indices grouped by channel group"""
    with get_db() as db:
        return [dict(row) for row in db.execute(
            """SELECT pi.*, cg.name as channel_group_name 
               FROM position_indices pi 
               JOIN channel_groups cg ON pi.channel_group_id = cg.id 
               ORDER BY cg.name, pi.position_type"""
        ).fetchall()]

def get_duration_index(channel_group, duration_seconds):
    """Get duration index for specific channel group and duration"""
    with get_db() as db:
        # First try to find by channel_group_id
        if isinstance(channel_group, int):
            channel_group_id = channel_group
        else:
            # Look up channel group ID by name
            cg_row = db.execute(
                "SELECT id FROM channel_groups WHERE name = ?", 
                (channel_group,)
            ).fetchone()
            channel_group_id = cg_row["id"] if cg_row else None
            
        if channel_group_id:
            row = db.execute(
                "SELECT index_value FROM duration_indices WHERE channel_group_id = ? AND duration_seconds = ?",
                (channel_group_id, duration_seconds)
            ).fetchone()
            return float(row["index_value"]) if row else 1.0
        
        return 1.0  # Default if no channel group found

def get_seasonal_index(channel_group, month):
    """Get seasonal index for specific channel group and month (1-12)"""
    with get_db() as db:
        # First try to find by channel_group_id
        if isinstance(channel_group, int):
            channel_group_id = channel_group
        else:
            # Look up channel group ID by name
            cg_row = db.execute(
                "SELECT id FROM channel_groups WHERE name = ?", 
                (channel_group,)
            ).fetchone()
            channel_group_id = cg_row["id"] if cg_row else None
            
        if channel_group_id:
            row = db.execute(
                "SELECT index_value FROM seasonal_indices WHERE channel_group_id = ? AND month = ?",
                (channel_group_id, month)
            ).fetchone()
            return float(row["index_value"]) if row else 1.0
        
        return 1.0  # Default if no channel group found

def get_position_index(channel_group, position_type):
    """Get position index for specific channel group and position type"""
    with get_db() as db:
        # First try to find by channel_group_id
        if isinstance(channel_group, int):
            channel_group_id = channel_group
        else:
            # Look up channel group ID by name
            cg_row = db.execute(
                "SELECT id FROM channel_groups WHERE name = ?", 
                (channel_group,)
            ).fetchone()
            channel_group_id = cg_row["id"] if cg_row else None
            
        if channel_group_id:
            row = db.execute(
                "SELECT index_value FROM position_indices WHERE channel_group_id = ? AND position_type = ?",
                (channel_group_id, position_type)
            ).fetchone()
            return float(row["index_value"]) if row else 1.0
        
        return 1.0  # Default if no channel group found

def update_duration_index(channel_group, duration_seconds, index_value, description=None):
    """Update or create duration index"""
    with get_db() as db:
        # Look up channel group ID by name if needed
        if isinstance(channel_group, str):
            cg_row = db.execute(
                "SELECT id FROM channel_groups WHERE name = ?", 
                (channel_group,)
            ).fetchone()
            channel_group_id = cg_row["id"] if cg_row else None
        else:
            channel_group_id = channel_group
            
        if channel_group_id:
            db.execute("""
                INSERT OR REPLACE INTO duration_indices (channel_group_id, duration_seconds, index_value, description)
                VALUES (?, ?, ?, ?)
            """, (channel_group_id, duration_seconds, index_value, description))
            db.commit()

def update_seasonal_index(channel_group, month, index_value, description=None):
    """Update seasonal index for specific channel group and month"""
    with get_db() as db:
        # Look up channel group ID by name if needed
        if isinstance(channel_group, str):
            cg_row = db.execute(
                "SELECT id FROM channel_groups WHERE name = ?", 
                (channel_group,)
            ).fetchone()
            channel_group_id = cg_row["id"] if cg_row else None
        else:
            channel_group_id = channel_group
            
        if channel_group_id:
            db.execute("""
                INSERT OR REPLACE INTO seasonal_indices (channel_group_id, month, index_value, description)
                VALUES (?, ?, ?, ?)
            """, (channel_group_id, month, index_value, description))
            db.commit()

def delete_duration_index(channel_group, duration_seconds):
    """Delete duration index"""
    with get_db() as db:
        # Look up channel group ID by name if needed
        if isinstance(channel_group, str):
            cg_row = db.execute(
                "SELECT id FROM channel_groups WHERE name = ?", 
                (channel_group,)
            ).fetchone()
            channel_group_id = cg_row["id"] if cg_row else None
        else:
            channel_group_id = channel_group
            
        if channel_group_id:
            db.execute("DELETE FROM duration_indices WHERE channel_group_id = ? AND duration_seconds = ?", 
                       (channel_group_id, duration_seconds))
            db.commit()

def get_target_groups_list():
    """Get all available target groups from TRP rates (indices are now by channel group)"""
    with get_db() as db:
        trp_tgs = db.execute("SELECT DISTINCT target_group FROM trp_rates").fetchall()
        
        # Get target groups from TRP rates only (indices are now channel group based)
        all_tgs = set()
        for row in trp_tgs:
            all_tgs.add(row["target_group"])
            
        return sorted(list(all_tgs))

def get_indices_for_wave_item(channel_group, duration_seconds, start_date, end_date=None):
    """Get appropriate duration and seasonal indices for wave item based on channel group"""
    duration_index = get_duration_index(channel_group, duration_seconds)
    
    # Calculate seasonal index based on wave date range
    seasonal_index = 1.0
    if start_date:
        try:
            from datetime import datetime
            start_obj = datetime.strptime(start_date, '%Y-%m-%d')
            
            # If end_date is provided and spans multiple months, calculate average
            if end_date:
                try:
                    end_obj = datetime.strptime(end_date, '%Y-%m-%d')
                    seasonal_index = calculate_average_seasonal_index(channel_group, start_obj, end_obj)
                    print(f"DEBUG: Multi-month wave {start_date} to {end_date}, average seasonal_index={seasonal_index}")
                except Exception as e:
                    print(f"DEBUG: Error parsing end_date {end_date}, using start_date only: {e}")
                    seasonal_index = get_seasonal_index(channel_group, start_obj.month)
            else:
                # Single month or no end date provided
                seasonal_index = get_seasonal_index(channel_group, start_obj.month)
                print(f"DEBUG: Single month wave {start_date}, seasonal_index={seasonal_index}")
                
        except Exception as e:
            print(f"DEBUG: Error parsing start_date {start_date}: {e}")
    else:
        print(f"DEBUG: No start_date provided for channel_group={channel_group}")
    
    return {
        'duration_index': duration_index,
        'seasonal_index': seasonal_index
    }

def calculate_average_seasonal_index(channel_group, start_date, end_date):
    """Calculate average seasonal index for a date range spanning multiple months"""
    from datetime import datetime, timedelta
    from calendar import monthrange
    
    if start_date.year != end_date.year or start_date.month == end_date.month:
        # Same month or different years, use simple approach
        return get_seasonal_index(channel_group, start_date.month)
    
    total_weighted_index = 0
    total_days = 0
    
    current_date = start_date
    while current_date <= end_date:
        # Calculate days in current month for this wave
        if current_date.month == start_date.month:
            # First month: from start_date to end of month
            _, days_in_month = monthrange(current_date.year, current_date.month)
            days_in_current_month = days_in_month - current_date.day + 1
        elif current_date.month == end_date.month:
            # Last month: from beginning of month to end_date
            days_in_current_month = end_date.day
        else:
            # Full middle months
            _, days_in_current_month = monthrange(current_date.year, current_date.month)
        
        # Get seasonal index for this month
        month_index = get_seasonal_index(channel_group, current_date.month)
        
        # Add to weighted total
        total_weighted_index += month_index * days_in_current_month
        total_days += days_in_current_month
        
        print(f"DEBUG: Month {current_date.month}, days={days_in_current_month}, index={month_index}")
        
        # Move to next month
        if current_date.month == 12:
            current_date = current_date.replace(year=current_date.year + 1, month=1, day=1)
        else:
            current_date = current_date.replace(month=current_date.month + 1, day=1)
    
    average_index = total_weighted_index / total_days if total_days > 0 else 1.0
    print(f"DEBUG: Average seasonal index calculation: total_weighted={total_weighted_index}, total_days={total_days}, average={average_index}")
    
    return average_index

def migrate_remove_pricing_list_requirement():
    """Remove pricing_list_id requirement from campaigns table"""
    with get_db() as db:
        # Check if pricing_list_id column exists and is NOT NULL
        cursor = db.execute("PRAGMA table_info(campaigns)")
        columns = {row[1]: row for row in cursor.fetchall()}
        
        if 'pricing_list_id' in columns:
            col_info = columns['pricing_list_id']
            is_not_null = col_info[3] == 1  # notnull flag
            
            if is_not_null:
                print("Removing NOT NULL constraint from campaigns.pricing_list_id...")
                
                # SQLite doesn't support ALTER COLUMN, so we need to recreate the table
                # First, get all existing data
                existing_campaigns = db.execute("SELECT * FROM campaigns").fetchall()
                
                # Drop the old table (after backing up foreign key constraints)
                db.execute("PRAGMA foreign_keys=OFF")
                db.execute("DROP TABLE campaigns")
                
                # Recreate campaigns table without NOT NULL on pricing_list_id
                db.execute("""
                CREATE TABLE campaigns (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    pricing_list_id INTEGER,  -- Removed NOT NULL
                    start_date TEXT,
                    end_date TEXT,
                    agency TEXT,
                    client TEXT, 
                    product TEXT,
                    country TEXT,
                    status TEXT
                )
                """)
                
                # Restore data, setting pricing_list_id to NULL where needed
                for row in existing_campaigns:
                    values = list(row)
                    # If pricing_list_id references non-existent pricing list, set to NULL
                    if len(values) > 2 and values[2]:  # pricing_list_id
                        try:
                            check = db.execute("SELECT 1 FROM pricing_lists WHERE id = ?", (values[2],)).fetchone()
                            if not check:
                                values[2] = None
                        except:
                            values[2] = None
                    
                    # Handle cases where row might not have all columns
                    while len(values) < 11:  # Pad with None for missing columns
                        values.append(None)
                    
                    db.execute("""
                    INSERT INTO campaigns (id, name, pricing_list_id, start_date, end_date, 
                                         agency, client, product, country, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, values[:10])
                
                db.execute("PRAGMA foreign_keys=ON")
                db.commit()
                print("Successfully removed NOT NULL constraint from pricing_list_id")
            else:
                print("pricing_list_id already allows NULL values")
        else:
            print("pricing_list_id column does not exist")

        db.commit()


def export_channel_group_excel(group_id: int):
    """Export Excel file for all campaigns using this channel group"""
    from datetime import datetime
    import sys
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    print(f"DEBUG: Starting Excel export for group_id={group_id}", file=sys.stderr, flush=True)

    # Quick test - return minimal Excel file
    if group_id == 996:
        print(f"DEBUG: Creating test Excel for group 996", file=sys.stderr, flush=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = f"Test Excel for group {group_id}"
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # Get channel group info
    group = get_channel_group_by_id(group_id)
    if not group:
        raise ValueError(f"Channel group {group_id} not found")

    group_name = group['name']
    print(f"DEBUG: Found group name={group_name}", file=sys.stderr)

    # Get all wave items that use channels from this group
    with get_db() as db:
        # First get the channel group name from the ID
        query = """
        SELECT wi.*, w.start_date, w.end_date, w.campaign_id, c.name as campaign_name,
               cg.name as channel_group_name, t.name as tvc_name,
               wi.duration_index, wi.seasonal_index,
               wi.trp_purchase_index, wi.advance_purchase_index, wi.web_index,
               wi.advance_payment_index, wi.loyalty_discount_index, wi.position_index
        FROM wave_items wi
        JOIN waves w ON wi.wave_id = w.id
        JOIN campaigns c ON w.campaign_id = c.id
        JOIN channel_groups cg ON cg.id = ?
        LEFT JOIN tvcs t ON wi.tvc_id = t.id
        WHERE wi.owner = cg.name
        ORDER BY c.name, w.start_date, wi.id
        """
        rows = db.execute(query, (group_id,)).fetchall()

    print(f"DEBUG: Found {len(rows)} rows for group_id={group_id}", file=sys.stderr, flush=True)

    # Load TRP distribution data for all campaigns using this channel group
    campaign_trp_data = {}
    if rows:
        campaign_ids = list(set(row['campaign_id'] for row in rows))
        print(f"DEBUG: Loading TRP data for campaigns: {campaign_ids}", file=sys.stderr, flush=True)
        for campaign_id in campaign_ids:
            print(f"DEBUG: Loading TRP data for campaign {campaign_id}", file=sys.stderr, flush=True)
            campaign_trp_data[campaign_id] = load_trp_distribution(campaign_id)
            print(f"DEBUG: Loaded TRP data for campaign {campaign_id}: {len(campaign_trp_data[campaign_id])} entries", file=sys.stderr, flush=True)

    if not rows:
        # Create empty Excel with message
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kanalų ataskaita"
        ws['A1'] = f"Kanalų grupės '{group_name}' planai nerasti"
        ws['A1'].font = Font(size=14, bold=True)
    else:
        # Create workbook with data
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kanalų ataskaita"

        # Styles
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))

        current_row = 1

        # Main header
        ws.merge_cells(f'A{current_row}:V{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"KANALŲ GRUPĖS '{group_name}' PLANŲ ATASKAITA"
        cell.font = Font(size=16, bold=True, color="1F4E79")
        cell.alignment = Alignment(horizontal='center')
        current_row += 2

        # Column headers - match the exact plan table columns
        headers = [
            'Pradžia', 'Pabaiga', 'Kanalų grupė', 'Kampanija', 'Perkama TG', 'TVC', 'Trukmė', 'TG\ndydis (*000)',
            'TG\ndalis (%)', 'TG\nimtis', 'Kanalo\ndalis', 'PT zonos\ndalis', 'nPT zonos\ndalis', 'GRP\nplanuojamas', 'TRP\nperkamas',
            'Affinity1', 'Gross CPP', 'Trukmės\nkoeficientas', 'Sezoninis\nkoeficientas', 'TRP\npirkimo',
            'Išankstinio\npirkimo', 'WEB', 'Išankstinio\nmokėjimo', 'Lojalumo\nnuolaida',
            'Gross\nkaina', 'Kliento\nnuolaida %', 'Net kaina', 'Agentūros\nnuolaida %', 'Net net kaina'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Set header row height to accommodate wrapped text
        ws.row_dimensions[current_row].height = 40

        current_row += 1

        # Store the starting row for data (needed for calendar alignment)
        data_start_row = current_row

        # Data rows
        print(f"DEBUG: About to process {len(rows)} rows", file=sys.stderr, flush=True)
        for item in rows:
            # Calculate values
            gross_cpp = item['gross_cpp_eur'] or 0
            grp_planned = (item['trps'] * 100 / item['affinity1']) if item['affinity1'] and item['affinity1'] > 0 else 0

            # Calculate base gross price (TRP * CPP * Duration) without multipliers
            # This shows the base calculation before all the coefficients/indices are applied
            base_gross_price = (item['trps'] or 0) * (item['gross_cpp_eur'] or 0) * (item['clip_duration'] or 0)
            print(f"DEBUG: TRP={item['trps']}, CPP={item['gross_cpp_eur']}, Duration={item['clip_duration']}, Base Gross={base_gross_price}", file=sys.stderr, flush=True)

            # Apply client discount to base gross price
            net_price = base_gross_price * (1 - (item['client_discount'] or 0) / 100)

            # Calculate net_net_price from net_price and agency discount
            net_net_price = net_price * (1 - (item['agency_discount'] or 0) / 100)

            # Row data matching the exact plan table columns
            row_data = [
                item['start_date'] or '',                                    # Pradžia
                item['end_date'] or '',                                      # Pabaiga
                item['channel_group_name'],                                  # Kanalų grupė
                item['campaign_name'],                                       # Kampanija
                item['target_group'],                                        # Perkama TG
                item['tvc_name'] or '',                                      # TVC
                item['clip_duration'] or 0,                                  # Trukmė
                item['tg_size_thousands'] or 0,                             # TG dydis (*000) - from db
                (item['tg_share_percent'] or 0) / 100 if item['tg_share_percent'] else 0,  # TG dalis (%) - from db, convert to decimal
                item['tg_sample_size'] or 0,                                # TG imtis - from db
                (item['channel_share'] or 0),                              # Kanalo dalis - already decimal
                (item['pt_zone_share'] or 0),                              # PT zonos dalis - already decimal
                (item['npt_zone_share'] or 0.45),                          # nPT zonos dalis - already decimal
                grp_planned,                                                 # GRP plan. (calculated)
                item['trps'] or 0,                                          # TRP perkamas
                item['affinity1'] or 0,                                     # Affinity1
                item['gross_cpp_eur'] or 0,                                 # Gross CPP
                item['duration_index'] or 1.0,                             # Trukmės koeficientas
                item['seasonal_index'] or 1.0,                             # Sezoninis koeficientas
                item['trp_purchase_index'] or 1.0,                         # TRP pirkimo
                item['advance_purchase_index'] or 1.0,                     # Išankstinio pirkimo
                item['web_index'] or 1.0,                                  # WEB
                item['advance_payment_index'] or 1.0,                      # Išankstinio mokėjimo
                item['loyalty_discount_index'] or 1.0,                     # Lojalumo nuolaida
                base_gross_price,                                           # Gross kaina
                item['client_discount'] or 0,                              # Kl. nuol. %
                net_price,                                                  # Net kaina
                item['agency_discount'] or 0,                              # Ag. nuol. %
                net_net_price                                               # Net net kaina
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = border

                # Add center alignment for Pradžia, Pabaiga, and Kanalų grupė columns
                if col in [1, 2, 3]:  # Pradžia, Pabaiga, Kanalų grupė
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Format numbers (column numbers shifted by 1 due to added Kampanija column)
                if col in [9, 11, 12, 13]:  # Percentage columns: TG dalis (%), Kanalo dalis, PT zonos dalis, nPT zonos dalis
                    cell.number_format = '0.00%'
                elif col in [18, 19, 20, 21, 22, 23, 24, 25]:  # Index columns
                    cell.number_format = '0.00'
                elif col in [14, 15, 17, 26, 28, 30]:  # Currency columns: GRP plan., TRP perkamas, Gross CPP, Gross kaina, Net kaina, Net net kaina
                    cell.number_format = '#,##0.00'
                elif col in [27, 29]:  # Discount percentage columns: Kl. nuol. %, Ag. nuol. %
                    cell.number_format = '0.0"%"'

            current_row += 1

        # Add totals row after all plan data
        if rows:
            # Calculate totals and averages
            total_grp = sum((item['trps'] * 100 / item['affinity1']) if item['affinity1'] and item['affinity1'] > 0 else 0 for item in rows)
            total_trp = sum(item['trps'] or 0 for item in rows)

            # Calculate average affinity (only from non-zero values)
            affinity_values = [item['affinity1'] for item in rows if item['affinity1'] and item['affinity1'] > 0]
            avg_affinity = sum(affinity_values) / len(affinity_values) if affinity_values else 0

            # Calculate total prices
            total_gross = 0
            total_net = 0
            total_net_net = 0

            for item in rows:
                # Calculate base gross price (TRP * CPP * Duration) without multipliers
                base_gross_price = (item['trps'] or 0) * (item['gross_cpp_eur'] or 0) * (item['clip_duration'] or 0)
                net_price = base_gross_price * (1 - (item['client_discount'] or 0) / 100)
                net_net_price = net_price * (1 - (item['agency_discount'] or 0) / 100)

                total_gross += base_gross_price
                total_net += net_price
                total_net_net += net_net_price

            # Add totals row
            totals_row_data = [''] * 29  # Create empty row with 29 columns (reduced by 1 after removing position index)
            totals_row_data[0] = 'VISO:'  # First column shows "VISO:"
            totals_row_data[13] = total_grp    # GRP plan. (column 14, index 13)
            totals_row_data[14] = total_trp    # TRP perkamas (column 15, index 14)
            totals_row_data[15] = avg_affinity # Affinity1 average (column 16, index 15)
            totals_row_data[24] = total_gross  # Gross kaina (column 25, index 24)
            totals_row_data[26] = total_net    # Net kaina (column 27, index 26)
            totals_row_data[28] = total_net_net # Net net kaina (column 29, index 28)

            # Style totals row
            total_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")  # Light orange
            total_font = Font(bold=True, size=10)

            for col, value in enumerate(totals_row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.font = total_font
                cell.fill = total_fill
                cell.border = border

                # Add center alignment for Pradžia, Pabaiga, and Kanalų grupė columns
                if col in [1, 2, 3]:  # Pradžia, Pabaiga, Kanalų grupė
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Apply number formatting for totals
                if col in [14, 15, 17, 26, 28, 30]:  # Currency/number columns
                    if value and isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                elif col == 16:  # Affinity average
                    if value and isinstance(value, (int, float)):
                        cell.number_format = '0.00'

            current_row += 1

        # Add calendar section to the right of the main table
        if rows:
            try:
                from datetime import datetime, timedelta
                import json

                # Find the date range from all wave items
                start_dates = [datetime.strptime(item['start_date'], '%Y-%m-%d') for item in rows if item['start_date']]
                end_dates = [datetime.strptime(item['end_date'], '%Y-%m-%d') for item in rows if item['end_date']]

                if start_dates and end_dates:
                    start_date = min(start_dates)
                    end_date = max(end_dates)

                    # Calendar positioning - main table now has 30 columns (A-AD), calendar starts further right for better spacing
                    calendar_start_col = 35  # Calendar data starts at AI (35) - push further to the right

                    # Calendar headers start at row 1 to align with main table
                    # Month headers at row 1, day numbers at row 2, weekdays at row 3
                    # Then calendar data starts at data_start_row (matches main data)
                    month_header_row = 1
                    calendar_end_col = min(calendar_start_col + (end_date - start_date).days + 5, 60)

                    # Month headers
                    current_date = start_date
                    col_idx = calendar_start_col
                    months = []
                    month_spans = {}

                    while current_date <= end_date:
                        month_year = current_date.strftime('%Y-%m')
                        if month_year not in month_spans:
                            month_spans[month_year] = {'start': col_idx, 'name': current_date.strftime('%B %Y')}
                            months.append(month_year)
                        month_spans[month_year]['end'] = col_idx
                        current_date += timedelta(days=1)
                        col_idx += 1

                    # Create month headers at row 1 - show month name only once per month (merged)
                    for month_year in months:
                        span = month_spans[month_year]
                        start_col = span['start']
                        end_col = span['end']

                        # Set value and styling for the first cell
                        month_cell = ws.cell(row=month_header_row, column=start_col)
                        month_cell.value = span['name']
                        month_cell.font = Font(bold=True, size=8, color="FFFFFF")  # White text
                        month_cell.alignment = Alignment(horizontal='center')
                        month_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Same dark blue as main headers
                        month_cell.border = border

                        # If spanning multiple columns, apply same styling to all cells before merging
                        if start_col < end_col:
                            # Apply styling to all cells in the span before merging
                            for col in range(start_col, end_col + 1):
                                cell = ws.cell(row=month_header_row, column=col)
                                cell.font = Font(bold=True, size=8, color="FFFFFF")
                                cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                                cell.border = border

                            # Now merge the cells
                            ws.merge_cells(f'{openpyxl.utils.get_column_letter(start_col)}{month_header_row}:{openpyxl.utils.get_column_letter(end_col)}{month_header_row}')
                        else:
                            # Single column month, just apply border
                            month_cell.border = border

                    # Day headers at row 2
                    current_date = start_date
                    col_idx = calendar_start_col
                    while current_date <= end_date:
                        day_cell = ws.cell(row=2, column=col_idx)
                        day_cell.value = current_date.strftime('%d')
                        day_cell.font = Font(size=9, bold=True)
                        day_cell.alignment = Alignment(horizontal='center')
                        day_cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                        day_cell.border = border
                        current_date += timedelta(days=1)
                        col_idx += 1

                    # Weekday headers at row 3
                    current_date = start_date
                    col_idx = calendar_start_col
                    while current_date <= end_date:
                        weekday_cell = ws.cell(row=3, column=col_idx)
                        weekday_names = ['Pr', 'An', 'Tr', 'Kt', 'Pn', 'Št', 'Sk']
                        weekday_cell.value = weekday_names[current_date.weekday()]
                        weekday_cell.font = Font(size=8)
                        weekday_cell.alignment = Alignment(horizontal='center')
                        weekday_cell.border = border
                        current_date += timedelta(days=1)
                        col_idx += 1

                    # No need for separate calendar headers since Kampanija is now in main table

                    # Add plan rows with TRP values - one row per plan (matching main table Y positions)
                    # Use the stored data_start_row to match exact positions
                    for plan_idx, item in enumerate(rows):
                        # Use the exact same Y position as the corresponding plan row in main table
                        row_idx = data_start_row + plan_idx

                        # No need for separate plan labels since campaign name is now in main table

                        # TRP values for each day for this specific plan
                        current_date = start_date
                        col_idx = calendar_start_col

                        while current_date <= end_date:
                            daily_trp = 0
                            date_str = current_date.strftime('%Y-%m-%d')

                            # Check if this specific plan is active on this date
                            if (item['start_date'] and item['end_date'] and
                                item['start_date'] <= date_str <= item['end_date']):

                                # Get TRP value from campaign TRP distribution data
                                campaign_id = item['campaign_id']
                                if campaign_id in campaign_trp_data:
                                    trp_data = campaign_trp_data[campaign_id]
                                    if date_str in trp_data:
                                        # Use actual TRP calendar data
                                        daily_trp = trp_data[date_str]
                                    else:
                                        # If no TRP calendar data for this date, distribute evenly across wave period
                                        wave_start = datetime.strptime(item['start_date'], '%Y-%m-%d')
                                        wave_end = datetime.strptime(item['end_date'], '%Y-%m-%d')
                                        wave_days = (wave_end - wave_start).days + 1
                                        daily_trp = (item['trps'] or 0) / wave_days if wave_days > 0 else 0
                                else:
                                    # No TRP distribution data for campaign, distribute evenly
                                    wave_start = datetime.strptime(item['start_date'], '%Y-%m-%d')
                                    wave_end = datetime.strptime(item['end_date'], '%Y-%m-%d')
                                    wave_days = (wave_end - wave_start).days + 1
                                    daily_trp = (item['trps'] or 0) / wave_days if wave_days > 0 else 0

                            # Create cell for EVERY day in the date range (whether active or not)
                            trp_cell = ws.cell(row=row_idx, column=col_idx)

                            if daily_trp > 0:
                                trp_cell.value = round(daily_trp, 1)
                                trp_cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                            else:
                                trp_cell.value = ""
                                trp_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                            trp_cell.font = Font(size=8)
                            trp_cell.alignment = Alignment(horizontal='center')
                            trp_cell.border = border

                            current_date += timedelta(days=1)
                            col_idx += 1

                    # No need for separate plan label column widths

                    # Set calendar date columns to narrow width (approximately 0.4cm)
                    total_days = (end_date - start_date).days + 1
                    for col in range(calendar_start_col, calendar_start_col + total_days):
                        col_letter = openpyxl.utils.get_column_letter(col)
                        ws.column_dimensions[col_letter].width = 5

            except Exception as e:
                # If calendar generation fails, skip it
                pass


        # Set specific widths for columns to properly display content
        ws.column_dimensions['A'].width = 12  # Pradžia
        ws.column_dimensions['B'].width = 12  # Pabaiga
        ws.column_dimensions['C'].width = 18  # Kanalų grupė
        ws.column_dimensions['D'].width = 15  # Kampanija
        ws.column_dimensions['E'].width = 10  # Perkama TG - thinner
        ws.column_dimensions['F'].width = 8   # TVC - thinner
        ws.column_dimensions['G'].width = 7   # Trukmė - thinner
        ws.column_dimensions['H'].width = 8   # TG dydis (*000) - thinner
        ws.column_dimensions['I'].width = 7   # TG dalis (%) - thinner
        ws.column_dimensions['J'].width = 7   # TG imtis - thinner
        ws.column_dimensions['K'].width = 7   # Kanalo dalis - thinner
        ws.column_dimensions['L'].width = 8   # PT zonos dalis - thinner
        ws.column_dimensions['M'].width = 9   # nPT zonos dalis - thinner
        ws.column_dimensions['N'].width = 9   # GRP planuojamas - thinner
        ws.column_dimensions['O'].width = 8   # TRP perkamas - thinner
        ws.column_dimensions['P'].width = 8   # Affinity1 - thinner
        ws.column_dimensions['Q'].width = 9   # Gross CPP - thinner
        ws.column_dimensions['R'].width = 9   # Trukmės koeficientas - thinner
        ws.column_dimensions['S'].width = 9   # Sezoninis koeficientas - thinner
        ws.column_dimensions['T'].width = 8   # TRP pirkimo - thinner
        ws.column_dimensions['U'].width = 9   # Išankstinio pirkimo - thinner
        ws.column_dimensions['V'].width = 6   # WEB - thinner
        ws.column_dimensions['W'].width = 9   # Išankstinio mokėjimo - thinner
        ws.column_dimensions['X'].width = 9   # Lojalumo nuolaida - thinner
        ws.column_dimensions['Y'].width = 9   # Gross kaina - thinner
        ws.column_dimensions['Z'].width = 9   # Kliento nuolaida % - thinner
        ws.column_dimensions['AA'].width = 9  # Net kaina - thinner
        ws.column_dimensions['AB'].width = 9  # Agentūros nuolaida % - thinner
        ws.column_dimensions['AC'].width = 10 # Net net kaina - thinner

        # Auto-adjust column widths if content is wider than preset widths
        try:
            print(f"DEBUG: Starting auto-adjust column widths", file=sys.stderr, flush=True)
            column_count = 0
            for column in ws.columns:
                column_count += 1
                if column_count > 100:  # Safety limit to prevent infinite loops
                    print(f"DEBUG: Breaking auto-adjust after 100 columns", file=sys.stderr, flush=True)
                    break

                max_length = 0
                column_letter = None
                for cell in column:
                    try:
                        # Skip merged cells
                        if hasattr(cell, 'column_letter'):
                            if column_letter is None:
                                column_letter = cell.column_letter
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                if column_letter:
                    col_index = openpyxl.utils.column_index_from_string(column_letter)
                    # Skip calendar columns (AE onwards)
                    if col_index >= 31:
                        continue
                    # Skip Pradžia and Pabaiga columns from auto-adjustment to keep them narrow
                    if column_letter in ['A', 'B']:
                        continue
                    # Get current width
                    current_width = ws.column_dimensions[column_letter].width
                    # Only adjust if content requires more width
                    if max_length + 2 > current_width:
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
            print(f"DEBUG: Finished auto-adjust, processed {column_count} columns", file=sys.stderr, flush=True)
        except Exception as e:
            print(f"DEBUG: Error in auto-adjust: {e}", file=sys.stderr, flush=True)

    # Save to BytesIO
    print(f"DEBUG: Saving workbook to BytesIO", file=sys.stderr, flush=True)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    print(f"DEBUG: Excel generation complete, returning buffer", file=sys.stderr, flush=True)
    return output
